# -*- coding: utf-8 -*-
"""
Gera OE05 preenchido (PT, sem emoji) a partir do template estrutural.
- Preenche inputs (azul) com valores VERIFICADOS do motor M6 (Base, hub_on, fim-2024, em k EUR)
  + inputs de mercado (Damodaran / atenção.txt).
- Remove todos os emoji: nomes de folhas, títulos, estados e referências em fórmulas.
- Guarda num ficheiro NOVO (não destrói o original).
"""
import io, sys, re
import openpyxl

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

SRC = r"C:\Users\amfmn\OneDrive - Universidade de Aveiro\PEF\M5🕐\OE5🕐\OE05_G18_260601.xlsx"
DST = r"C:\Users\amfmn\OneDrive - Universidade de Aveiro\PEF\M5🕐\OE5🕐\OE05_G18_260601_PREENCHIDO.xlsx"

# ------------------------------------------------------------------ emoji
EMOJI_RE = re.compile(
    "["
    "\U0001F000-\U0001FAFF"   # pictogramas, simbolos, etc. (rosto, dinheiro, etc.)
    "\U00002600-\U000027BF"   # simbolos diversos + dingbats (engrenagem incl.)
    "\U0000FE00-\U0000FE0F"   # seletores de variacao
    "]+"
)
def strip_emoji(s: str) -> str:
    out = EMOJI_RE.sub("", s)
    out = re.sub(r"[ \t]{2,}", " ", out)   # colapsa espacos duplos deixados pelo emoji
    return out.strip()

wb = openpyxl.load_workbook(SRC)

# mapa nome-antigo -> nome-novo (sem emoji)
rename = {}
for ws in wb.worksheets:
    rename[ws.title] = strip_emoji(ws.title)

# ------------------------------------------------------------------ valores M6 + mercado
# Folha Pressupostos (coluna C). Numeros do motor M6 em k EUR; mercado = Damodaran.
pressupostos = {
    "C6":  "Ceramica / Bens de Casa (Household Products - Europa)",
    "C8":  2024,            # ano base
    "C10": 5,               # horizonte explicito (anos)
    "C12": 0.028,           # Rf
    "C13": 0.0543,          # ERP
    "C14": "=0.77*(1+(1-C16)*(C19/C20))",  # beta alavancado (re-alavancado c/ D/E Grestel)
    "C16": 0.21,            # taxa de imposto (IRC estatutario)
    "C17": 0.028,           # kd bruto
    "C19": 17733.8,         # divida financeira fim-2024
    "C20": 12199.7,         # capital proprio fim-2024 (total_cp)
    "C25": 0.08,            # g1 (descritivo)
    "C26": 0.062,           # g2 (descritivo)
    "C27": 0.0225,          # gn terminal
    "C29": 0.1221,          # ROC
    "C31": 1981.0,          # EBIT 2024
    "C32": 2168.7,          # D&A 2024
    "C33": 1224.7,          # CapEx 2024
    "C34": 5562.4,          # dNWC 2024 (referencia)
    "C36": 17191.4,         # divida liquida fim-2024
    # C37 (n.o accoes) deixado em branco -> preencher c/ R&C
    "C39": 15.86,           # EV/EBITDA (ja presente)
    "C40": 19.00,           # EV/EBIT
    "C41": 23.20,           # P/E
    "C42": 4.73,            # P/BV
    "C43": 2.79,            # EV/Sales
}

# Folha DCF-FCFF (colunas C..G = 2025..2029)
COLS = ["C", "D", "E", "F", "G"]
dcf_rows = {
    9:  [42051.0, 44816.4, 47676.6, 50834.1, 53815.2],   # Receita
    11: [5918.9, 7270.5, 8733.6, 9716.2, 10126.6],        # EBITDA
    13: [2108.3, 2827.3, 2741.7, 2614.4, 2405.3],         # D&A
    21: [4500.0, 4770.0, 1300.0, 1080.0, 860.0],          # CapEx
    22: [-214.7, -1159.7, 495.8, 662.8, 757.2],           # dNWC
}

# Folha FCFE (colunas C..G = 2025..2029)
fcfe_rows = {
    9:  [3061.6, 3348.9, 4544.5, 5475.6, 6065.2],         # Lucro Liquido
    10: [2391.7, 1942.7, -1441.7, -1534.4, -1545.3],      # CapEx Liquido
    12: [-591.2, -1604.6, -4038.6, -2173.9, -2012.9],     # Nova Divida Liquida
}

# Folha Sintese: pesos + desconto de negociacao
sintese = {
    "D4": 0.50,    # peso DCF
    "D5": 0.30,    # peso Multiplos
    "D6": 0.20,    # peso FCFE
    "C11": -0.10,  # desconto de negociacao (DLOM)
}

# ------------------------------------------------------------------ aplicar valores (pelos nomes ANTIGOS)
ws_p = wb["⚙️ Pressupostos"]
for coord, val in pressupostos.items():
    ws_p[coord] = val

ws_d = wb["📊 DCF-FCFF"]
for row, vals in dcf_rows.items():
    for col, v in zip(COLS, vals):
        ws_d[f"{col}{row}"] = v

ws_f = wb["💰 FCFE"]
for row, vals in fcfe_rows.items():
    for col, v in zip(COLS, vals):
        ws_f[f"{col}{row}"] = v

ws_s = wb["🏆 Síntese"]
for coord, val in sintese.items():
    ws_s[coord] = val

# ------------------------------------------------------------------ remover emoji de TODO o conteudo
# 1) fórmulas: trocar referencias de folha (com emoji) pelas novas (sem emoji)
# 2) texto: tirar emoji
for ws in wb.worksheets:
    for row in ws.iter_rows():
        for c in row:
            v = c.value
            if not isinstance(v, str):
                continue
            if v.startswith("="):
                nv = v
                for old, new in rename.items():
                    nv = nv.replace(f"'{old}'", f"'{new}'")
                if nv != v:
                    c.value = nv
            else:
                nv = strip_emoji(v)
                if nv != v:
                    c.value = nv

# atualizar estados no Indice (apos strip ficam 'Preencher' -> 'Preenchido')
ws_i = wb["📋 Índice"]
for coord in ("D6", "D7", "D9"):
    if ws_i[coord].value and "Preencher" in str(ws_i[coord].value):
        ws_i[coord] = "Preenchido (M6)"

# ------------------------------------------------------------------ renomear folhas (por ultimo)
for ws in wb.worksheets:
    ws.title = rename[ws.title]

wb.save(DST)
print("OK guardado em:", DST)
print("Folhas:", wb.sheetnames)
