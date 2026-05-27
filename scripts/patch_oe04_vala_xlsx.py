"""Patch OE04_G18_260525_VALA.xlsx — align VALA escudo with MSD sheets and WACC 6,3%."""
from __future__ import annotations

import sys
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell

DOC_PATH = Path(__file__).resolve().parents[1] / "OE04_G18_260525_VALA.xlsx"


def set_cell(ws, row: int, col: int, value) -> None:
    """Write to top-left cell if target is inside a merged range."""
    for merged in ws.merged_cells.ranges:
        if (
            merged.min_row <= row <= merged.max_row
            and merged.min_col <= col <= merged.max_col
        ):
            ws.cell(merged.min_row, merged.min_col).value = value
            return
    cell = ws.cell(row, col)
    if not isinstance(cell, MergedCell):
        cell.value = value

# 11_VALA: cols C..L = years 2025..2034 (col 3..12)
YEAR_COLS = list(range(3, 13))
# 05_MSD_BancoHub: row 16 = 2025 .. 25 = 2034, col F = juros expensed
BH_ROW = {2025 + i: 16 + i for i in range(10)}
# 06_MSD_BEI: row 13 = 2025 .. 22 = 2034, col F = juros expensed
BEI_ROW = {2025 + i: 13 + i for i in range(10)}


def col_letter(col: int) -> str:
    from openpyxl.utils import get_column_letter

    return get_column_letter(col)


def patch_11_vala(ws) -> None:
    t_ref = "$C$10"
    kd_bh = "$C$6"
    kd_bei = "$C$7"
    rf_ref = "$C$9"
    for i, col in enumerate(YEAR_COLS):
        year = 2025 + i
        cl = col_letter(col)
        bh_r = BH_ROW[year]
        bei_r = BEI_ROW[year]
        exp_n = f"COLUMN({cl}$29)-COLUMN($C$29)+1"

        # Row 30 — Juros expensed Banco Hub
        set_cell(ws, 30, col, f"=IFERROR('05_MSD_BancoHub'!F{bh_r},0)")
        set_cell(ws, 31, col, f"=IFERROR('06_MSD_BEI'!F{bei_r},0)")
        set_cell(ws, 32, col, f"={cl}30+{cl}31")
        set_cell(ws, 33, col, f"=ROUND({cl}30*{t_ref},0)")
        set_cell(ws, 34, col, f"=ROUND({cl}31*{t_ref},0)")
        set_cell(ws, 35, col, f"={cl}33+{cl}34")
        set_cell(
            ws,
            36,
            col,
            f"=IF({cl}30=0,0,ROUND({cl}33/((1+{kd_bh})^({exp_n})),0))",
        )
        set_cell(
            ws,
            37,
            col,
            f"=IF({cl}31=0,0,ROUND({cl}34/((1+{kd_bei})^({exp_n})),0))",
        )

    set_cell(ws, 59, 4, "=SUM(C36:L36)+SUM(C37:L37)")

    # PT2030 accrual fiscal cost — link to FCF [3a] row 29
    for i, col in enumerate(YEAR_COLS):
        cl = col_letter(col)
        fcf_col = col_letter(col)  # same column in sheet 10 row 29
        set_cell(
            ws,
            44,
            col,
            f"=IF('10_FCF_Viabilidade'!{fcf_col}29=0,0,"
            f"-ROUND('10_FCF_Viabilidade'!{fcf_col}29*{t_ref},0))",
        )
        set_cell(ws, 45, col, f"={cl}43+{cl}44")
        exp_n = f"COLUMN({cl}$29)-COLUMN($C$29)+1"
        set_cell(
            ws,
            46,
            col,
            f"=IF({cl}45=0,0,ROUND({cl}45/((1+{rf_ref})^({exp_n})),0))",
        )

    set_cell(ws, 48, 3, "=SUM(C46:L46)")  # C48 merged with D48

    for i, col in enumerate(YEAR_COLS):
        cl = col_letter(col)
        fcf_col = col_letter(col)
        set_cell(ws, 52, col, f"=IFERROR('10_FCF_Viabilidade'!{fcf_col}34,0)")
        exp_n = f"COLUMN({cl}$29)-COLUMN($C$29)+1"
        set_cell(
            ws,
            53,
            col,
            f"=IF({cl}52=0,0,ROUND({cl}52/((1+{rf_ref})^({exp_n})),0))",
        )
    set_cell(ws, 55, 3, "=SUM(C53:L53)")

    set_cell(ws, 11, 3, "='00_Pressupostos'!B39")
    set_cell(ws, 11, 4, "we×Ke + wd×Kd×(1−T) = 6,30% (Folha 10)")
    set_cell(ws, 65, 6, "a WACC = 6,3% (Folha 10)")

    # Nota metodológica VALA (cell A69 or append)
    note = (
        "Nota: VALA é complemento analítico M6. O VAL oficial do projeto (Folha 10) "
        "usa WACC 6,3%. O escudo fiscal da dívida (secção C) reporta juros expensed "
        "das Folhas 05 e 06 — após amortização antecipada PT2030, juros Banco Hub = 0 desde 2029."
    )
    for r in range(65, 80):
        if ws.cell(r, 1).value and "Ver relatório" in str(ws.cell(r, 1).value):
            ws.cell(r, 1).value = note
            break


def patch_pressupostos(ws) -> None:
    label = ws.cell(48, 1).value
    if label and "7,3" in str(label):
        ws.cell(48, 1).value = str(label).replace("7,3%", "6,3%").replace("7.3%", "6.3%")


def patch_00a_synthesis(ws) -> None:
    """Optional: ensure synthesis references correct WACC if present."""
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=8):
        for cell in row:
            if cell.value and isinstance(cell.value, str) and "7,30%" in cell.value:
                cell.value = cell.value.replace("7,30%", "6,30%")


def main() -> int:
    path = Path(sys.argv[1]) if len(sys.argv) > 1 else DOC_PATH
    wb = load_workbook(str(path))
    patch_11_vala(wb["11_VALA"])
    patch_pressupostos(wb["00_Pressupostos"])
    if "00A_Síntese_OE4" in wb.sheetnames:
        patch_00a_synthesis(wb["00A_Síntese_OE4"])
    wb.save(str(path))
    print(f"Patched: {path}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
