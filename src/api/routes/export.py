"""Rota de exportação para Excel."""

import io
from datetime import datetime

from fastapi import APIRouter, Query
from fastapi.responses import StreamingResponse
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from src.engine.modelo.model import dataframe_to_records, run_model
from src.engine.modelo.pressupostos import build_pressupostos_summary
from src.engine.projetos.hub_logistico import (
    load as hub_load,
    viabilidade_hub,
    mapa_servico_divida,
)

router = APIRouter(prefix="/api")

_HDR_FILL = PatternFill("solid", fgColor="3B2A1A")
_HDR_FONT = Font(bold=True, color="FFFFFF", size=10)
_ALT_FILL = PatternFill("solid", fgColor="F7F0E8")
_SEC_FONT = Font(bold=True, color="3B2A1A", size=10)


def _write_header(ws, headers, row=1):
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font = _HDR_FONT
        cell.fill = _HDR_FILL
        cell.alignment = Alignment(horizontal="left")


def _write_records(ws, records, start_row=2):
    for i, row in enumerate(records):
        fill = _ALT_FILL if i % 2 == 0 else None
        for col, val in enumerate(row.values(), 1):
            cell = ws.cell(row=start_row + i, column=col, value=val)
            if fill:
                cell.fill = fill


def _sheet_from_records(ws, records):
    if not records:
        ws.cell(1, 1, "Sem dados")
        return
    _write_header(ws, list(records[0].keys()))
    _write_records(ws, records)
    _autofit(ws)


def _autofit(ws):
    for col_cells in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col_cells), default=8)
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(max_len + 3, 45)


@router.get("/export/excel")
def export_excel(
    cenario: str = Query("Base"),
    hub_on: bool = Query(False),
    ecogres_on: bool = Query(True),
):
    """Gera ficheiro Excel com todos os dados calculados pelo modelo financeiro."""
    dfs = run_model(cenario=cenario, hub_on=hub_on, ecogres_on=ecogres_on)
    rec = dataframe_to_records(dfs)
    pressupostos = build_pressupostos_summary(cenario=cenario, hub_on=hub_on, ecogres_on=ecogres_on)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # DR
    ws = wb.create_sheet("DR")
    _sheet_from_records(ws, rec.get("dr", []))

    # Balanço
    ws = wb.create_sheet("Balanço")
    _sheet_from_records(ws, rec.get("balanco", []))

    # DFC
    ws = wb.create_sheet("DFC")
    _sheet_from_records(ws, rec.get("dfc", []))

    # KPIs
    ws = wb.create_sheet("KPIs")
    _sheet_from_records(ws, rec.get("kpis", []))

    # FSE
    ws = wb.create_sheet("FSE")
    _sheet_from_records(ws, rec.get("fse_detalhe_anual", []))

    # Pessoal
    ws = wb.create_sheet("Pessoal")
    _sheet_from_records(ws, rec.get("pessoal_contab_anual", []))

    # Produção
    ws = wb.create_sheet("Produção")
    _sheet_from_records(ws, rec.get("producao_anual", []))

    # Pressupostos — flat: Secção | Parâmetro | Valor | Unidade | Nota
    ws = wb.create_sheet("Pressupostos")
    _write_header(ws, ["Secção", "Parâmetro", "Valor", "Unidade", "Nota"])
    row_idx = 2
    for section in pressupostos.get("sections", []):
        sec_label = section.get("label", "")
        for item in section.get("items", []):
            fill = _ALT_FILL if row_idx % 2 == 0 else None
            vals = [sec_label, item.get("label"), item.get("value"), item.get("unit", ""), item.get("note", "")]
            for col, val in enumerate(vals, 1):
                cell = ws.cell(row=row_idx, column=col, value=val)
                if fill:
                    cell.fill = fill
            row_idx += 1
    _autofit(ws)

    # Hub Logístico (apenas se ativo)
    if hub_on:
        hub = hub_load()
        hub_res = viabilidade_hub(hub)

        # Hub · Viabilidade
        ws_hv = wb.create_sheet("Hub_Viabilidade")
        kv_rows = [
            {"Indicador": "VAL (€)", "Valor": hub_res.get("val")},
            {"Indicador": "TIR", "Valor": hub_res.get("tir")},
            {"Indicador": "Payback Simples (anos)", "Valor": hub_res.get("payback_simples")},
            {"Indicador": "Payback Atualizado (anos)", "Valor": hub_res.get("payback_atualizado")},
            {"Indicador": "Índice de Rendibilidade", "Valor": hub_res.get("indice_rendibilidade")},
            {"Indicador": "Valor Terminal (€)", "Valor": hub_res.get("valor_terminal")},
            {"Indicador": "Valor Residual Ativos (€)", "Valor": hub_res.get("valor_residual_ativos")},
            {"Indicador": "NFM Recovery Terminal (€)", "Valor": hub_res.get("nfm_recovery_terminal")},
            {"Indicador": "Capital Vivo T10 (€)", "Valor": hub_res.get("capital_vivo_t10")},
            {"Indicador": "Mais-Valia (€)", "Valor": hub_res.get("mais_valia")},
        ]
        for k, v in (hub_res.get("parametros") or {}).items():
            kv_rows.append({"Indicador": f"Parâmetro: {k}", "Valor": v})
        _write_header(ws_hv, ["Indicador", "Valor"])
        _write_records(ws_hv, kv_rows)

        # FCF por ano — em bloco abaixo dos indicadores
        fcf_df = hub_res.get("fcf_df")
        if fcf_df is not None and hasattr(fcf_df, "to_dict"):
            fcf_records = fcf_df.to_dict(orient="records")
            if fcf_records:
                start = len(kv_rows) + 3
                ws_hv.cell(start, 1, "Free Cash Flow por Ano").font = _SEC_FONT
                _write_header(ws_hv, list(fcf_records[0].keys()), row=start + 1)
                _write_records(ws_hv, fcf_records, start_row=start + 2)
        _autofit(ws_hv)

        # Hub · Serviço da Dívida
        ws_hd = wb.create_sheet("Hub_Divida")
        debt_df = mapa_servico_divida(hub)
        if hasattr(debt_df, "to_dict"):
            _sheet_from_records(ws_hd, debt_df.to_dict(orient="records"))

    # Info
    ws_i = wb.create_sheet("Info")
    _sheet_from_records(ws_i, [
        {"Parâmetro": "Cenário", "Valor": cenario},
        {"Parâmetro": "Hub Logístico", "Valor": "Ativo" if hub_on else "Desativado"},
        {"Parâmetro": "Ecogres Consolidada", "Valor": "Sim" if ecogres_on else "Não"},
        {"Parâmetro": "Data de exportação", "Valor": datetime.now().strftime("%Y-%m-%d %H:%M")},
        {"Parâmetro": "Modelo", "Valor": "GrestelPy v0.9.5"},
    ])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M")
    filename = f"GrestelPy_{cenario}{'_Hub' if hub_on else ''}_{timestamp}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


# ─── M3 Export ────────────────────────────────────────────────────────────────

_ALL_YEARS = [2024, 2025, 2026, 2027, 2028, 2029]

_FMT = {
    "eur":   '#,##0 €',
    "eur2":  '#,##0.00 €',
    "pct":   '0.0%',
    "dias":  '0.0',
    "ratio": '0.00',
    "m_eur": '0.000',
}

_TITLE_FONT  = Font(bold=True, color="3B2A1A", size=11)
_YEAR_FILL   = PatternFill("solid", fgColor="3B2A1A")
_YEAR_FONT   = Font(bold=True, color="FFFFFF", size=10)
_SEP_FILL    = PatternFill("solid", fgColor="D9CFC4")
_SEP_FONT    = Font(bold=True, color="3B2A1A", size=10)


def _write_pivoted(ws, title: str, rows_config: list, data: dict, years: list, row_start: int = 1) -> int:
    """Escreve uma tabela pivotada: métricas nas linhas, anos nas colunas.

    rows_config: lista de (label, field, tipo) ou ("---", None, None) para separador.
    data: dict {ano: {field: valor}}.
    Retorna a próxima linha livre.
    """
    # Título
    ws.cell(row_start, 1, title).font = _TITLE_FONT
    r = row_start + 1

    # Header
    ws.cell(r, 1, "Rubrica").font = _HDR_FONT
    ws.cell(r, 1).fill = _HDR_FILL
    for ci, y in enumerate(years, 2):
        c = ws.cell(r, ci, str(y))
        c.font = _YEAR_FONT
        c.fill = _YEAR_FILL
        c.alignment = Alignment(horizontal="center")
    r += 1

    # Dados
    for i, (label, field, tipo) in enumerate(rows_config):
        if label == "---":          # separador visual
            for ci in range(1, len(years) + 2):
                c = ws.cell(r, ci)
                c.fill = _SEP_FILL
            r += 1
            continue

        fill = _ALT_FILL if i % 2 == 0 else None
        lbl_cell = ws.cell(r, 1, label)
        if fill:
            lbl_cell.fill = fill

        for ci, y in enumerate(years, 2):
            val = data.get(y, {}).get(field)
            cell = ws.cell(r, ci, val)
            if val is not None and tipo in _FMT:
                cell.number_format = _FMT[tipo]
            if fill:
                cell.fill = fill
            cell.alignment = Alignment(horizontal="right")
        r += 1

    _autofit(ws)
    return r + 1


def _compute_m3(dfs: dict) -> dict:
    """Extrai e deriva todas as métricas necessárias para as 8 tabelas M3.

    Retorna {ano: {campo: valor}} para anos 2024–2029.
    """
    df_dr  = dfs.get("dr",  None)
    df_bal = dfs.get("balanco", None)
    df_kpi = dfs.get("kpis", None)
    df_pes = dfs.get("pessoal_anual", None)

    out = {}
    for y in _ALL_YEARS:
        d = {}

        # ── DR ──────────────────────────────────────────────────────────────
        if df_dr is not None:
            row = df_dr[df_dr.ano == y]
            if not row.empty:
                r = row.iloc[0]
                d["vn"]              = float(r["vn"])
                d["outros_rend"]     = float(r.get("outros_rend", 0))
                d["cmvmc"]           = float(r["cmvmc"])          # negativo
                d["fse"]             = float(r["fse"])            # negativo
                d["gastos_pessoal"]  = float(r["gastos_pessoal"]) # negativo
                d["ebitda"]          = float(r["ebitda"])
                d["depreciacoes"]    = float(r["depreciacoes"])   # negativo
                d["ebit"]            = float(r["ebit"])
                d["juros"]           = float(r["juros"])          # negativo
                d["rai"]             = float(r["rai"])
                d["irc"]             = float(r["irc"])            # negativo
                d["rl"]              = float(r["rl"])

                vn = d["vn"] if d["vn"] else 1.0
                d["margem_bruta_eur"] = vn + d["cmvmc"]          # vn + cmvmc (cmvmc é neg)
                d["margem_bruta_pct"] = d["margem_bruta_eur"] / vn
                d["vbp"]              = vn + abs(d["outros_rend"])
                d["vab"]              = vn + abs(d["outros_rend"]) - abs(d["cmvmc"]) - abs(d["fse"])
                d["dep_abs"]          = abs(d["depreciacoes"])
                d["juros_abs"]        = abs(d["juros"])

                # % VN
                d["cmvmc_vn"]    = abs(d["cmvmc"])        / vn
                d["pessoal_vn"]  = abs(d["gastos_pessoal"]) / vn
                d["fse_vn"]      = abs(d["fse"])           / vn
                d["dep_vn"]      = d["dep_abs"]            / vn
                d["ebitda_margin"] = d["ebitda"]           / vn

                # % VBP
                vbp = d["vbp"] if d["vbp"] else 1.0
                for field in ["vn", "outros_rend", "vbp", "cmvmc", "margem_bruta_eur",
                               "gastos_pessoal", "fse", "ebitda", "dep_abs",
                               "ebit", "juros", "rai", "irc", "rl"]:
                    val = d.get(field, 0.0)
                    d[f"{field}_pct_vbp"] = val / vbp

        # ── Balanço ─────────────────────────────────────────────────────────
        if df_bal is not None:
            row = df_bal[df_bal.ano == y]
            if not row.empty:
                b = row.iloc[0]
                d["caixa"]        = float(b.get("caixa", 0))
                d["fornecedores"] = float(b.get("fornecedores", 0))
                d["eoep_credor"]  = float(b.get("eoep_credor", 0))
                d["outros_pc"]    = float(b.get("outros_pc", 0))
                d["emprestimos_c"]  = float(b.get("emprestimos_c", 0))
                d["linha_credito"]  = float(b.get("linha_credito_cp", 0))
                d["passivo_corrente"] = (
                    d["fornecedores"] + d["eoep_credor"] + d["outros_pc"]
                    + d["emprestimos_c"] + d["linha_credito"]
                )
                d["liquidez_imediata"] = (
                    d["caixa"] / d["passivo_corrente"]
                    if d["passivo_corrente"] else 0.0
                )

        # ── KPIs ────────────────────────────────────────────────────────────
        if df_kpi is not None:
            row = df_kpi[df_kpi.ano == y]
            if not row.empty:
                k = row.iloc[0]
                d["roa"]                 = float(k["roa"])
                d["roe"]                 = float(k["roe"])
                d["rl_margin"]           = float(k["rl_margin"])      # ROS
                d["autonomia_financeira"]= float(k["autonomia_financeira"])
                d["solvabilidade"]       = float(k["solvabilidade"])
                d["liquidez_geral"]      = float(k["liquidez_geral"])
                d["liquidez_reduzida"]   = float(k["liquidez_reduzida"])
                d["debt_equity"]         = float(k["debt_equity"])
                d["dscr"]                = float(k["dscr"])
                d["pmr_dias"]            = float(k["pmr_dias"])
                d["pmp_dias"]            = float(k["pmp_dias"])
                d["dmi_dias"]            = float(k["dmi_dias"])
                d["ciclo_caixa"]         = float(k["ciclo_caixa"])
                d["cresc_vn"]            = float(k.get("cresc_vn", 0))
                d["juros_abs_kpi"]       = float(k.get("juros_abs", 0))
                d["ebitda_kpi"]          = float(k.get("ebitda", d.get("ebitda", 0)))
                d["cobertura_ebitda_juros"] = (
                    d["ebitda_kpi"] / d["juros_abs_kpi"]
                    if d["juros_abs_kpi"] else 0.0
                )

        # ── Pessoal ─────────────────────────────────────────────────────────
        if df_pes is not None:
            row = df_pes[df_pes.ano == y]
            if not row.empty:
                hc = float(row.iloc[0].get("headcount", 1))
                vn = d.get("vn", 0)
                vab = d.get("vab", 0)
                d["headcount"]          = hc
                d["vab_por_colaborador"]= vab / hc if hc else 0.0
                d["vn_por_colaborador"] = vn  / hc if hc else 0.0

        # m€ convenience
        for field in ["vn", "ebitda", "ebit", "rl"]:
            if field in d:
                d[f"{field}_meur"] = d[field] / 1_000_000

        out[y] = d

    return out


@router.get("/export/m3")
def export_m3(
    cenario: str = Query("Base"),
    hub_on: bool = Query(False),
    ecogres_on: bool = Query(True),
):
    """Excel M3: 8 sheets alinhadas às tabelas do relatório académico M3 (secções 7.1–13.3)."""
    dfs  = run_model(cenario=cenario, hub_on=hub_on, ecogres_on=ecogres_on)
    data = _compute_m3(dfs)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    years = _ALL_YEARS

    # ── Sheet 1: 7.1 DR Orçamental (2024 Hist vs 2025 Orç + Var) ─────────────
    ws = wb.create_sheet("7.1 DR Orçamental")
    hdrs = ["Rubrica", "2024 (Hist.)", "2025 (Orç.)", "Var %", "Var abs (€)"]
    for ci, h in enumerate(hdrs, 1):
        c = ws.cell(1, ci, h)
        c.font = _HDR_FONT; c.fill = _HDR_FILL
        c.alignment = Alignment(horizontal="left" if ci == 1 else "center")

    rows_71 = [
        ("Vendas e Prestações de Serviços", "vn",            "eur"),
        ("CMVMC",                           "cmvmc",         "eur"),
        ("Margem Bruta",                    "margem_bruta_eur","eur"),
        ("Gastos com Pessoal",              "gastos_pessoal","eur"),
        ("FSE",                             "fse",           "eur"),
        ("EBITDA",                          "ebitda",        "eur"),
        ("Depreciações e Amortizações",     "dep_abs",       "eur"),
        ("EBIT",                            "ebit",          "eur"),
        ("Gastos de Financiamento",         "juros",         "eur"),
        ("RAI",                             "rai",           "eur"),
        ("IRC",                             "irc",           "eur"),
        ("Resultado Líquido",               "rl",            "eur"),
    ]
    d24, d25 = data.get(2024, {}), data.get(2025, {})
    for i, (label, field, tipo) in enumerate(rows_71):
        r = i + 2
        fill = _ALT_FILL if i % 2 == 0 else None
        v24 = d24.get(field)
        v25 = d25.get(field)
        var_pct = ((v25 / v24) - 1) if (v24 and v24 != 0) else None
        var_abs = (v25 - v24)        if (v24 is not None and v25 is not None) else None

        for ci, (val, fmt_str) in enumerate([
            (label,   None),
            (v24,     _FMT["eur"]),
            (v25,     _FMT["eur"]),
            (var_pct, _FMT["pct"]),
            (var_abs, _FMT["eur"]),
        ], 1):
            cell = ws.cell(r, ci, val)
            if fmt_str:
                cell.number_format = fmt_str
            if fill:
                cell.fill = fill
            if ci > 1:
                cell.alignment = Alignment(horizontal="right")
    _autofit(ws)

    # ── Sheet 2: 10.2 Estrutura de Gastos (% VN) ─────────────────────────────
    ws = wb.create_sheet("10.2 Estrutura Gastos")
    _write_pivoted(ws, "Estrutura de Gastos Operacionais (% VN) — Tabela 10.2", [
        ("CMVMC / VN",                "cmvmc_vn",    "pct"),
        ("Gastos com Pessoal / VN",   "pessoal_vn",  "pct"),
        ("FSE / VN",                  "fse_vn",      "pct"),
        ("D&A / VN",                  "dep_vn",      "pct"),
        ("Margem EBITDA",             "ebitda_margin","pct"),
    ], data, years)

    # ── Sheet 3: 10.3 Resultados Projetados ───────────────────────────────────
    ws = wb.create_sheet("10.3 Resultados")
    _write_pivoted(ws, "Resultados Projetados — Tabela 10.3", [
        ("VN (M€)",               "vn_meur",    "m_eur"),
        ("EBITDA (M€)",           "ebitda_meur","m_eur"),
        ("EBIT (M€)",             "ebit_meur",  "m_eur"),
        ("Resultado Líquido (M€)","rl_meur",    "m_eur"),
        ("ROE (%)",               "roe",        "pct"),
    ], data, years)

    # ── Sheet 4: 11.1 DR Comparativa (€ + % VBP) ─────────────────────────────
    ws = wb.create_sheet("11.1 DR Comparativa")
    rows_eur = [
        ("Vendas e PS",                   "vn",               "eur"),
        ("Outros Rendimentos Operacionais","outros_rend",      "eur"),
        ("VBP",                           "vbp",              "eur"),
        ("CMVMC",                         "cmvmc",            "eur"),
        ("Margem Bruta",                  "margem_bruta_eur", "eur"),
        ("Gastos com Pessoal",            "gastos_pessoal",   "eur"),
        ("FSE",                           "fse",              "eur"),
        ("EBITDA",                        "ebitda",           "eur"),
        ("Depreciações e Amortizações",   "dep_abs",          "eur"),
        ("EBIT",                          "ebit",             "eur"),
        ("Gastos de Financiamento",       "juros",            "eur"),
        ("RAI",                           "rai",              "eur"),
        ("IRC",                           "irc",              "eur"),
        ("Resultado Líquido",             "rl",               "eur"),
    ]
    rows_pct = [
        (f"{lbl} (% VBP)", f"{field}_pct_vbp", "pct")
        for lbl, field, _ in rows_eur
    ]
    r_next = _write_pivoted(ws, "DR Comparativa — Valores em € — Tabela 11.1", rows_eur, data, years)
    _write_pivoted(ws, "DR Comparativa — Peso relativo no VBP — Tabela 11.1", rows_pct, data, years, row_start=r_next)

    # ── Sheet 5: 11.5 Indicadores Financeiros ────────────────────────────────
    ws = wb.create_sheet("11.5 Indicadores")
    _write_pivoted(ws, "Indicadores Financeiros Consolidados — Tabela 11.5", [
        ("Autonomia Financeira (%)",       "autonomia_financeira",    "pct"),
        ("Solvabilidade (%)",              "solvabilidade",           "pct"),
        ("Liquidez Geral",                 "liquidez_geral",          "ratio"),
        ("ROE (%)",                        "roe",                     "pct"),
        ("ROA (%)",                        "roa",                     "pct"),
        ("ROS — Margem Líquida (%)",       "rl_margin",               "pct"),
        ("Cobertura do Serviço da Dívida", "dscr",                    "ratio"),
        ("EBITDA / Encargos Financeiros",  "cobertura_ebitda_juros",  "ratio"),
    ], data, years)

    # ── Sheet 6: 13.1 Rácios Económicos ──────────────────────────────────────
    ws = wb.create_sheet("13.1 Rácios Econ.")
    _write_pivoted(ws, "Rácios Económicos — Tabela 13.1", [
        ("ROS — Result. Operac./VN (%)",  "rl_margin",           "pct"),
        ("ROA — RAI/AT (%)",              "roa",                 "pct"),
        ("ROE — RL/CP (%)",               "roe",                 "pct"),
        ("Margem EBITDA (%)",             "ebitda_margin",       "pct"),
        ("Margem Bruta (%)",              "margem_bruta_pct",    "pct"),
        ("VAB / Colaborador (€)",         "vab_por_colaborador", "eur"),
        ("VN / Colaborador (€)",          "vn_por_colaborador",  "eur"),
    ], data, years)

    # ── Sheet 7: 13.2 Rácios Financeiros ─────────────────────────────────────
    ws = wb.create_sheet("13.2 Rácios Fin.")
    _write_pivoted(ws, "Rácios Financeiros — Tabela 13.2", [
        ("Autonomia Financeira (%)",       "autonomia_financeira", "pct"),
        ("Solvabilidade (%)",              "solvabilidade",        "pct"),
        ("Liquidez Geral",                 "liquidez_geral",       "ratio"),
        ("Liquidez Reduzida",              "liquidez_reduzida",    "ratio"),
        ("Liquidez Imediata",              "liquidez_imediata",    "ratio"),
        ("Cobertura Serviço Dívida (DSCR)","dscr",                "ratio"),
        ("Debt / Equity",                  "debt_equity",          "ratio"),
    ], data, years)

    # ── Sheet 8: 13.3 Rácios de Atividade ────────────────────────────────────
    ws = wb.create_sheet("13.3 Rácios Ativ.")
    _write_pivoted(ws, "Rácios de Atividade — Tabela 13.3", [
        ("PMR — Prazo Médio de Recebimento (dias)", "pmr_dias",    "dias"),
        ("PMP — Prazo Médio de Pagamento (dias)",   "pmp_dias",    "dias"),
        ("PMI — Inventários (dias)",                "dmi_dias",    "dias"),
        ("Ciclo de Conversão de Caixa (dias)",      "ciclo_caixa", "dias"),
    ], data, years)

    # ── Sheet: Pressupostos ───────────────────────────────────────────────────
    pressupostos = build_pressupostos_summary(cenario=cenario, hub_on=hub_on, ecogres_on=ecogres_on)
    ws_p = wb.create_sheet("Pressupostos")
    _write_header(ws_p, ["Secção", "Parâmetro", "Valor", "Unidade", "Nota"])
    row_idx = 2
    for section in pressupostos.get("sections", []):
        sec_label = section.get("label", "")
        for item in section.get("items", []):
            fill = _ALT_FILL if row_idx % 2 == 0 else None
            vals = [sec_label, item.get("label"), item.get("value"), item.get("unit", ""), item.get("note", "")]
            for col, val in enumerate(vals, 1):
                cell = ws_p.cell(row=row_idx, column=col, value=val)
                if fill:
                    cell.fill = fill
            row_idx += 1
    _autofit(ws_p)

    # ── Sheet: Mixes (calculados pelo modelo) ─────────────────────────────────
    ws_m = wb.create_sheet("Mixes")
    df_vprod = dfs.get("vendas_produto_anual")
    df_vmerc = dfs.get("vendas_mercadoria_anual")
    df_vmkt  = dfs.get("vendas_mercado_anual")

    def _mix_row(ws, r, vals_fmts, fill):
        for col, (val, fmt) in enumerate(vals_fmts, 1):
            cell = ws.cell(r, col, val)
            if fmt:
                cell.number_format = fmt
                cell.alignment = Alignment(horizontal="right")
            if fill:
                cell.fill = fill

    # Tabela 1 — Mix por Produto (2024 vs 2025)
    ws_m.cell(1, 1, "Mix por Produto (% VN Produtos) — modelo").font = _TITLE_FONT
    _write_header(ws_m, ["Produto", "VN 2024 (€)", "Mix 2024", "VN 2025 (€)", "Mix 2025"], row=2)
    if df_vprod is not None and not df_vprod.empty:
        vn_24p = df_vprod[df_vprod.ano == 2024].groupby("produto")["vn"].sum()
        vn_25p = df_vprod[df_vprod.ano == 2025].groupby("produto")["vn"].sum()
        tot24p = float(vn_24p.sum()) or 1.0
        tot25p = float(vn_25p.sum()) or 1.0
        produtos_list = sorted(df_vprod["produto"].unique())
        for i, prod in enumerate(produtos_list):
            r = 3 + i
            v24 = float(vn_24p.get(prod, 0.0))
            v25 = float(vn_25p.get(prod, 0.0))
            _mix_row(ws_m, r, [
                (prod,           None),
                (v24,            _FMT["eur"]),
                (v24 / tot24p,   _FMT["pct"]),
                (v25,            _FMT["eur"]),
                (v25 / tot25p,   _FMT["pct"]),
            ], _ALT_FILL if i % 2 == 0 else None)
        row_mkt_tbl = 3 + len(produtos_list) + 2
    else:
        row_mkt_tbl = 5

    # Tabela 2 — Mix por Mercado Geográfico (2024 vs 2025)
    ws_m.cell(row_mkt_tbl, 1, "Mix por Mercado Geográfico (% VN Total) — modelo").font = _TITLE_FONT
    _write_header(ws_m, ["Mercado", "VN 2024 (€)", "Mix 2024", "VN 2025 (€)", "Mix 2025"], row=row_mkt_tbl + 1)
    if df_vmkt is not None and not df_vmkt.empty:
        vmkt_24 = df_vmkt[df_vmkt.ano == 2024].set_index("mercado")
        vmkt_25 = df_vmkt[df_vmkt.ano == 2025].set_index("mercado")
        mercados_list = sorted(df_vmkt["mercado"].unique())
        for i, mkt in enumerate(mercados_list):
            r = row_mkt_tbl + 2 + i
            v24 = float(vmkt_24.at[mkt, "vn"]) if mkt in vmkt_24.index else 0.0
            p24 = float(vmkt_24.at[mkt, "peso"]) if mkt in vmkt_24.index else 0.0
            v25 = float(vmkt_25.at[mkt, "vn"]) if mkt in vmkt_25.index else 0.0
            p25 = float(vmkt_25.at[mkt, "peso"]) if mkt in vmkt_25.index else 0.0
            _mix_row(ws_m, r, [
                (mkt,  None),
                (v24,  _FMT["eur"]),
                (p24,  _FMT["pct"]),
                (v25,  _FMT["eur"]),
                (p25,  _FMT["pct"]),
            ], _ALT_FILL if i % 2 == 0 else None)
        row_merc_tbl = row_mkt_tbl + 2 + len(mercados_list) + 2
    else:
        row_merc_tbl = row_mkt_tbl + 4

    # Tabela 3 — Mix por Mercadoria (2024 vs 2025)
    ws_m.cell(row_merc_tbl, 1, "Mix por Mercadoria (% VN Mercadorias) — modelo").font = _TITLE_FONT
    _write_header(ws_m, ["Mercadoria", "VN 2024 (€)", "Mix 2024", "VN 2025 (€)", "Mix 2025"], row=row_merc_tbl + 1)
    if df_vmerc is not None and not df_vmerc.empty:
        vmerc_24 = df_vmerc[df_vmerc.ano == 2024].set_index("mercadoria")["vn"]
        vmerc_25 = df_vmerc[df_vmerc.ano == 2025].set_index("mercadoria")["vn"]
        tot_m24 = float(vmerc_24.sum()) or 1.0
        tot_m25 = float(vmerc_25.sum()) or 1.0
        mercadorias_list = sorted(df_vmerc["mercadoria"].unique())
        for i, merc in enumerate(mercadorias_list):
            r = row_merc_tbl + 2 + i
            v24 = float(vmerc_24.get(merc, 0.0))
            v25 = float(vmerc_25.get(merc, 0.0))
            _mix_row(ws_m, r, [
                (merc,           None),
                (v24,            _FMT["eur"]),
                (v24 / tot_m24,  _FMT["pct"]),
                (v25,            _FMT["eur"]),
                (v25 / tot_m25,  _FMT["pct"]),
            ], _ALT_FILL if i % 2 == 0 else None)
    _autofit(ws_m)

    # ── Sheet: PVU (todos os anos, produtos e mercadorias) ────────────────────
    ws_pvu = wb.create_sheet("PVU")
    df_prod_anual = dfs.get("producao_anual")

    # PVU médio por produto (ponderado por qtd através de todos os mercados)
    ws_pvu.cell(1, 1, "PVU Médio por Produto (€/un.) — calculado pelo modelo").font = _TITLE_FONT
    _write_header(ws_pvu, ["Produto"] + [str(y) for y in years], row=2)
    if df_prod_anual is not None and not df_prod_anual.empty:
        produtos_pvu = sorted(df_prod_anual["produto"].unique())
        for i, prod in enumerate(produtos_pvu):
            r = 3 + i
            fill = _ALT_FILL if i % 2 == 0 else None
            cell_l = ws_pvu.cell(r, 1, prod)
            if fill:
                cell_l.fill = fill
            for ci, y in enumerate(years, 2):
                mask = (df_prod_anual["ano"] == y) & (df_prod_anual["produto"] == prod)
                pvu_val = float(df_prod_anual.loc[mask, "pvu"].iloc[0]) if mask.any() else None
                cell = ws_pvu.cell(r, ci, pvu_val)
                if pvu_val is not None:
                    cell.number_format = _FMT["eur2"]
                cell.alignment = Alignment(horizontal="right")
                if fill:
                    cell.fill = fill
        row_merc_pvu = 3 + len(produtos_pvu) + 2
    else:
        row_merc_pvu = 5

    # PVU por mercadoria
    ws_pvu.cell(row_merc_pvu, 1, "PVU por Mercadoria (€/un.) — calculado pelo modelo").font = _TITLE_FONT
    _write_header(ws_pvu, ["Mercadoria"] + [str(y) for y in years], row=row_merc_pvu + 1)
    if df_vmerc is not None and not df_vmerc.empty:
        mercadorias_pvu = sorted(df_vmerc["mercadoria"].unique())
        for i, merc in enumerate(mercadorias_pvu):
            r = row_merc_pvu + 2 + i
            fill = _ALT_FILL if i % 2 == 0 else None
            cell_l = ws_pvu.cell(r, 1, merc)
            if fill:
                cell_l.fill = fill
            for ci, y in enumerate(years, 2):
                mask = (df_vmerc["ano"] == y) & (df_vmerc["mercadoria"] == merc)
                pvu_val = float(df_vmerc.loc[mask, "pvu"].iloc[0]) if mask.any() else None
                cell = ws_pvu.cell(r, ci, pvu_val)
                if pvu_val is not None:
                    cell.number_format = _FMT["eur2"]
                cell.alignment = Alignment(horizontal="right")
                if fill:
                    cell.fill = fill
    _autofit(ws_pvu)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M")
    filename = f"GrestelPy_M3_{cenario}_{timestamp}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )
