"""Gera a OE05 PREENCHIDA (10 anos, 2025-2034) reproduzindo o motor M6.

Pipeline de transcrição uni-direccional (memória OE5): corre o motor M6
(run_model + VALA), extrai os números e ESCREVE-os na folha de cálculo da OE5.
O código NÃO lê a OE5 — só a produz. Reproduz exactamente a fonte de verdade:

  - DCF-FCFF 10 anos, valor terminal de Gordon SIMPLES (FCFF₂₀₃₄×(1+gₙ)/(WACC−gₙ)),
    gₙ = 2,0 % — igual a GrestelModel._equity_dcf.
  - FCFE pela ponte FCFF→equity (FCFF − Juros×(1−t) + ΔDívida) — igual ao motor.
  - Múltiplos = MÉDIA dos 5 (EV/EBITDA, EV/EBIT, EV/Sales, P/E, P/BV) — igual ao motor.
  - WACC = fórmula viva com benefício fiscal; pesos a valor de mercado (equity VALA).
  - Síntese por hierarquia de fiabilidade: intrínseco pesa o preço (DCF-FCFF 60 % +
    FCFE 40 %); múltiplos = banda de confronto (0 %). Desconto −10 % — reproduz min_price.

Resultado-alvo (cenário Base): DCF 63 577 k€, FCFE 49 892 k€ (intrínseco ponderado
58 103 k€); Múltiplos 71 053 k€ só como confronto; preço mínimo ≈ 99,36 €/acção.
"""
from __future__ import annotations

import os

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from src.api.routes.valuation import _GRESTEL_DEFAULTS, run_model
from src.engine.valuation import compute_vala

# ── Paleta e estilos (iguais ao template estrutural) ──────────────────────────
C_INPUT = "1F4FD0"
C_REF = "1F8A4C"
C_HEAD_BG = "1F3A5F"
C_SECT_BG = "DCE6F1"
C_TITLE_BG = "2E5496"

F_TITLE = Font(name="Calibri", size=13, bold=True, color="FFFFFF")
F_HEAD = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
F_SECT = Font(name="Calibri", size=10, bold=True, color="1F3A5F")
F_INPUT = Font(name="Calibri", size=10, color=C_INPUT, bold=True)
F_REF = Font(name="Calibri", size=10, color=C_REF)
F_CALC = Font(name="Calibri", size=10, color="000000")
F_LABEL = Font(name="Calibri", size=10, color="000000")
F_NOTE = Font(name="Calibri", size=8, italic=True, color="666666")

FILL_HEAD = PatternFill("solid", fgColor=C_HEAD_BG)
FILL_SECT = PatternFill("solid", fgColor=C_SECT_BG)
FILL_TITLE = PatternFill("solid", fgColor=C_TITLE_BG)

_thin = Side(style="thin", color="BFBFBF")
BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)

PCT = "0.00%"
PCT1 = "0.0%"
EURMIL = "#,##0"
EUR2 = "#,##0.00"
MULT = '0.00"x"'
FACT = "0.0000"

ALIGN_L = Alignment(horizontal="left", vertical="center", wrap_text=False)
ALIGN_R = Alignment(horizontal="right", vertical="center")
ALIGN_C = Alignment(horizontal="center", vertical="center", wrap_text=True)

# Nomes de folha (planos, iguais ao ficheiro PREENCHIDO existente)
S_IDX = "Índice"
S_PARAMS = "Pressupostos"
S_DCF = "DCF-FCFF"
S_MULT = "Múltiplos"
S_FCFE = "FCFE"
S_STRESS = "Stress Tests"
S_SYN = "Síntese"

YEARS = list(range(2025, 2035))           # 2025..2034 (10 anos)
NY = len(YEARS)
YCOL = [get_column_letter(3 + i) for i in range(NY)]   # C..L
LAST = YCOL[-1]                                          # L


# ── Helpers de escrita ────────────────────────────────────────────────────────
def title(ws, text, span):
    ws.merge_cells(start_row=1, start_column=2, end_row=1, end_column=1 + span)
    c = ws.cell(1, 2, text)
    c.font = F_TITLE; c.fill = FILL_TITLE; c.alignment = ALIGN_L
    ws.row_dimensions[1].height = 22


def label(ws, r, text, *, col=2, bold=False, note=False):
    c = ws.cell(r, col, text)
    c.font = F_NOTE if note else (F_SECT if bold else F_LABEL)
    c.alignment = ALIGN_L
    return c


def section(ws, r, text, span):
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=1 + span)
    c = ws.cell(r, 2, text)
    c.font = F_SECT; c.fill = FILL_SECT; c.alignment = ALIGN_L


def head(ws, r, cols):
    for col, text in cols:
        c = ws.cell(r, col, text)
        c.font = F_HEAD; c.fill = FILL_HEAD; c.alignment = ALIGN_C; c.border = BORDER


def put_in(ws, r, c, value, *, fmt=None, unit=None, source=None):
    cell = ws.cell(r, c, value)
    cell.font = F_INPUT; cell.alignment = ALIGN_R; cell.border = BORDER
    if fmt: cell.number_format = fmt
    if unit is not None: ws.cell(r, c + 1, unit).font = F_NOTE
    if source is not None: ws.cell(r, c + 2, source).font = F_NOTE
    return cell


def put_f(ws, r, c, formula, *, fmt=None, ref=False, unit=None, source=None):
    cell = ws.cell(r, c, formula)
    cell.font = F_REF if ref else F_CALC
    cell.alignment = ALIGN_R; cell.border = BORDER
    if fmt: cell.number_format = fmt
    if unit is not None: ws.cell(r, c + 1, unit).font = F_NOTE
    if source is not None: ws.cell(r, c + 2, source).font = F_NOTE
    return cell


# ── Extracção dos números do motor M6 ─────────────────────────────────────────
def extract_engine(cenario="Base"):
    p = dict(_GRESTEL_DEFAULTS)
    tax = float(p["tax_rate"])
    dfs = run_model(cenario=cenario, hub_on=False, ecogres_on=True,
                    horizonte_maturidade=True)
    dr, dfc, bal = dfs["dr"], dfs["dfc"], dfs["balanco"]

    def grow(df):  # dict ano->row
        return {int(r["ano"]): r for _, r in df.iterrows()}
    drm, dfcm, balm = grow(dr), grow(dfc), grow(bal)

    def debt(y):
        b = balm.get(y)
        if b is None: return 0.0
        return (float(b.get("emprestimos_nc", 0) or 0)
                + float(b.get("emprestimos_c", 0) or 0)
                + float(b.get("linha_credito_cp", 0) or 0))

    rows = []
    for y in YEARS:
        d = drm[y]; f = dfcm[y]
        vn = float(d.get("vn", 0) or 0)
        ebitda = float(d.get("ebitda", 0) or 0)
        da = -float(d.get("depreciacoes", 0) or 0)          # add-back (>0)
        ebit = float(d.get("ebit", 0) or 0)
        juros = abs(float(d.get("juros", 0) or 0))
        capex = (abs(float(f.get("capex_aft", 0) or 0))
                 + abs(float(f.get("pag_intang", 0) or 0))
                 + abs(float(f.get("hub_capex", 0) or 0)))
        dnwc = -float(f.get("var_nfm", 0) or 0)             # ΔNWC = −var_nfm
        net_borrow = debt(y) - debt(y - 1)
        nopat = ebit * (1 - tax)
        fcff = nopat + da - capex - dnwc
        fcfe = fcff - juros * (1 - tax) + net_borrow
        rows.append(dict(ano=y, vn=vn, ebitda=ebitda, da=da, ebit=ebit,
                         capex=capex, dnwc=dnwc, juros=juros,
                         net_borrow=net_borrow, nopat=nopat, fcff=fcff, fcfe=fcfe))

    net_debt = debt(2024) - (float(balm[2024].get("caixa", 0) or 0)
                             + float(balm[2024].get("aplicacoes_fin_cp", 0) or 0))
    book_eq = float(balm[2024].get("total_cp", 0) or 0)

    vala = compute_vala(
        beta_u=float(p["beta_u"]), rf=float(p["rf"]), erp=float(p["erp"]),
        kd=float(p["kd"]), tax_rate=tax, net_debt=net_debt,
        fcffs=[r["fcff"] for r in rows], g_terminal=float(p["g_terminal"]),
    )
    return dict(rows=rows, net_debt=net_debt, book_eq=book_eq, vala=vala,
                tax=tax, defaults=p)


K = 1000.0  # € → € mil


# ── Folhas ────────────────────────────────────────────────────────────────────
def build_idx(ws):
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 64
    ws.column_dimensions["D"].width = 18
    title(ws, "OE05 — MODELO DE AVALIAÇÃO · GRESTEL (reproduz motor M6, 10 anos)", span=3)
    label(ws, 2, "Grestel — Produtos Cerâmicos, S.A.   ·   PEF 2025-26   ·   Grupo 18")
    head(ws, 4, [(2, "Folha"), (3, "Conteúdo"), (4, "Estado")])
    rows = [
        (S_IDX, "Navegação e legenda", "—"),
        (S_PARAMS, "Custo de capital (WACC c/ benefício fiscal), crescimento, múltiplos", "Motor M6"),
        (S_DCF, "DCF-FCFF (obrigatório) — 10 anos 2025-2034, Gordon simples", "Motor M6"),
        (S_MULT, "Avaliação relativa — média de 5 múltiplos (EV/EBITDA, EV/EBIT, EV/Sales, P/E, P/BV)", "Ligado"),
        (S_FCFE, "Ótica do accionista — ponte FCFF→equity, descontado a ke", "Motor M6"),
        (S_STRESS, "Sensibilidade dinâmica WACC × gₙ (fórmula viva, 10 anos)", "Automático"),
        (S_SYN, "Comparação dos 3 métodos + preço mínimo de negociação", "Automático"),
    ]
    r = 5
    for nome, cont, estado in rows:
        label(ws, r, nome)
        ws.cell(r, 3, cont).font = F_LABEL
        ws.cell(r, 4, estado).alignment = ALIGN_C
        for cc in (2, 3, 4): ws.cell(r, cc).border = BORDER
        r += 1
    r += 1
    section(ws, r, "LEGENDA", span=3); r += 1
    for txt, fnt in [
        ("● Azul — input transcrito do motor M6 (fonte de verdade)", F_INPUT),
        ("● Verde — referência a outra folha", F_REF),
        ("● Preto — fórmula calculada", F_CALC),
    ]:
        ws.cell(r, 2, txt).font = fnt; r += 1
    r += 1
    label(ws, r, "Os números são transcritos do motor M6 (run_model + VALA, cenário Base, "
                 "horizonte de maturidade 2025-2034). O Excel calcula tudo com fórmulas próprias.", note=True)


def build_params(ws, E):
    for col, w in {"B": 42, "C": 16, "D": 10, "E": 64}.items():
        ws.column_dimensions[col].width = w
    p = E["defaults"]; vala = E["vala"]
    mkt_eq = vala.equity_vala / K
    title(ws, "PRESSUPOSTOS E PARÂMETROS DO MODELO", span=4)
    label(ws, 2, "Inputs a AZUL ← motor M6. Fórmulas (preto) calculam automaticamente.", note=True)
    head(ws, 3, [(2, "Parâmetro"), (3, "Valor"), (4, "Unid."), (5, "Fonte / Notas")])

    section(ws, 4, "A. EMPRESA", span=4)
    label(ws, 5, "Nome da Empresa"); ws.cell(5, 3, "Grestel — Produtos Cerâmicos, S.A.").font = F_INPUT
    ws.cell(5, 5, "Identificação").font = F_NOTE
    label(ws, 6, "Setor / Indústria"); ws.cell(6, 3, "Cerâmica / Household Products (Europa)").font = F_INPUT
    ws.cell(6, 5, "← Damodaran: Household Products (Europe)").font = F_NOTE
    label(ws, 7, "Moeda de Reporte"); ws.cell(7, 4, "EUR").font = F_NOTE
    label(ws, 8, "Ano Base (histórico)"); put_in(ws, 8, 3, 2024, unit="Ano", source="Último ano real (R&C) ← M6")
    label(ws, 9, "Primeiro Ano de Projeção"); put_f(ws, 9, 3, "=C8+1", unit="Ano")
    label(ws, 10, "Horizonte de Projeção (anos)")
    put_in(ws, 10, 3, NY, unit="Anos", source="2025-2034 (horizonte de maturidade do motor)")

    section(ws, 11, "B. CUSTO DE CAPITAL (WACC)", span=4)
    label(ws, 12, "Taxa Livre de Risco (Rf)")
    put_in(ws, 12, 3, float(p["rf"]), fmt=PCT, unit="%", source="← OT PT 10a (BPstat 2025)")
    label(ws, 13, "Prémio de Risco de Mercado (ERP)")
    put_in(ws, 13, 3, float(p["erp"]), fmt=PCT, unit="%", source="← Damodaran ERP Portugal (5-jan-2026)")
    label(ws, 14, "Beta Alavancado (βL)")
    put_in(ws, 14, 3, round(vala.beta_l, 6), fmt="0.00", unit="—",
           source="← M6: βU=0,71 (sector) re-alavancado Hamada c/ D/E=0,28")
    label(ws, 15, "Custo Capital Próprio (ke)")
    put_f(ws, 15, 3, "=C12+C14*C13", fmt=PCT, unit="%",
          source="CAPM: ke = Rf + βL×ERP (fórmula viva). ke ALAVANCADO entra no WACC. FCFF/WACC: NÃO usa Ku.")
    label(ws, 16, "Taxa de Imposto (t)")
    put_in(ws, 16, 3, float(p["tax_rate"]), fmt=PCT, unit="%", source="IRC vigente ← M6")
    label(ws, 17, "Custo Dívida Bruta (kd)")
    put_in(ws, 17, 3, float(p["kd"]), fmt=PCT, unit="%", source="← M6: Rf + spread de crédito")
    label(ws, 18, "Custo Dívida Após Impostos")
    put_f(ws, 18, 3, "=C17*(1-C16)", fmt=PCT, unit="%", source="kd×(1−t)  ← BENEFÍCIO FISCAL")
    label(ws, 19, "Dívida Líquida (pesos WACC)")
    put_f(ws, 19, 3, "=C36", fmt=EURMIL, unit="€ mil", source="= Dívida líquida — CONCEITO ÚNICO (=C36)")
    label(ws, 20, "Capital Próprio (valor de mercado)")
    put_in(ws, 20, 3, round(mkt_eq, 1), fmt=EURMIL, unit="€ mil",
           source="← M6: equity intrínseco (VALA) — peso-alvo do WACC a valor de mercado (Damodaran)")
    label(ws, 21, "Peso Dívida (D/V)"); put_f(ws, 21, 3, '=IFERROR(C19/(C19+C20),"")', fmt=PCT, unit="%")
    label(ws, 22, "Peso Capital Próprio (E/V)"); put_f(ws, 22, 3, '=IFERROR(C20/(C19+C20),"")', fmt=PCT, unit="%")
    label(ws, 23, "WACC", bold=True)
    put_f(ws, 23, 3, "=C15*C22+C18*C21", fmt=PCT, unit="%",
          source="WACC = ke×E/V + kd×(1−t)×D/V — fórmula viva c/ BENEFÍCIO FISCAL (C18)")

    section(ws, 24, "C. CRESCIMENTO E VALOR TERMINAL", span=4)
    label(ws, 25, "Crescimento Terminal (gₙ)")
    put_in(ws, 25, 3, float(p["g_terminal"]), fmt=PCT, unit="%",
           source="≤ crescimento nominal da economia (Damodaran cap.12) ← M6")
    label(ws, 26, "Método do Valor Terminal")
    ws.cell(26, 3, "Gordon simples").font = F_INPUT
    ws.cell(26, 5, "FCFF₂₀₃₄×(1+gₙ)/(WACC−gₙ) — igual a GrestelModel._equity_dcf").font = F_NOTE
    label(ws, 27, "Nota crescimento", note=True)
    ws.cell(27, 5, "As taxas de crescimento estão embebidas nas séries projetadas (← M6), não recalculadas aqui.").font = F_NOTE

    section(ws, 30, "D. DADOS FINANCEIROS (€ mil)", span=4)
    r0 = E["rows"][0]
    label(ws, 31, "EBIT 2025E"); put_in(ws, 31, 3, round(r0["ebit"] / K, 1), fmt=EURMIL, unit="€ mil", source="← M6: DR 2025")
    label(ws, 32, "Depreciações & Amortizações 2025E"); put_in(ws, 32, 3, round(r0["da"] / K, 1), fmt=EURMIL, unit="€ mil", source="← M6")
    label(ws, 33, "CapEx 2025E"); put_in(ws, 33, 3, round(r0["capex"] / K, 1), fmt=EURMIL, unit="€ mil", source="← M6: DFC")
    label(ws, 34, "EBITDA 2025E"); put_f(ws, 34, 3, "=C31+C32", fmt=EURMIL, unit="€ mil", source="= EBIT + D&A")
    label(ws, 35, "Capital Próprio Contabilístico"); put_in(ws, 35, 3, round(E["book_eq"] / K, 1), fmt=EURMIL, unit="€ mil", source="← M6: total CP 2024 (base do P/BV)")
    label(ws, 36, "Dívida Líquida", bold=True)
    put_in(ws, 36, 3, round(E["net_debt"] / K, 4), fmt=EURMIL, unit="€ mil", source="← M6: dívida fin. − caixa 2024 (CONCEITO ÚNICO)")
    label(ws, 37, "Nº de Ações em Circulação", bold=True)
    put_in(ws, 37, 3, int(p["shares"]), fmt="#,##0", unit="unid.", source="← R&C: capital social 526 318 € a valor nominal 1 €")

    section(ws, 38, "E. MÚLTIPLOS DE REFERÊNCIA (SETOR, Damodaran)", span=4)
    for r, nome, key in [
        (39, "EV/EBITDA — Setor", "EV_EBITDA_sector"),
        (40, "EV/EBIT — Setor", "EV_EBIT_sector"),
        (41, "P/E — Setor", "PE_sector"),
        (42, "P/BV — Setor", "PBV_sector"),
        (43, "EV/Sales — Setor", "EV_Sales_sector"),
    ]:
        label(ws, r, nome)
        put_in(ws, r, 3, float(p[key]), fmt=MULT, unit="x", source="← Mercado: Damodaran (setor europeu)")


def build_dcf(ws, E):
    ws.column_dimensions["B"].width = 40
    for col in YCOL: ws.column_dimensions[col].width = 11
    rows = E["rows"]
    title(ws, "MODELO DCF — FREE CASH FLOW TO THE FIRM (FCFF) · 10 anos", span=1 + NY)
    label(ws, 2, "Séries (azul) ← motor M6. Valor terminal de Gordon simples. Valuation = fórmulas.", note=True)
    head(ws, 3, [(2, "Item")] + [(3 + i, f"Ano {i+1}\n{y}") for i, y in enumerate(YEARS)])

    section(ws, 4, "A. PRESSUPOSTOS (ligados a Pressupostos)", span=1 + NY)
    label(ws, 6, "WACC (desconto)")
    label(ws, 7, "Taxa de Imposto (t)")
    for i, col in enumerate(YCOL):
        put_f(ws, 6, 3 + i, f"='{S_PARAMS}'!$C$23", fmt=PCT, ref=True)
        put_f(ws, 7, 3 + i, f"='{S_PARAMS}'!$C$16", fmt=PCT, ref=True)

    section(ws, 8, "B. PROJEÇÃO DA DEMONSTRAÇÃO DE RESULTADOS (€ mil)", span=1 + NY)
    label(ws, 9, "Receita (Volume de Negócios)")
    label(ws, 10, "  Crescimento YoY", note=True)
    label(ws, 11, "EBITDA")
    label(ws, 12, "  Margem EBITDA", note=True)
    label(ws, 13, "Depreciações (D&A)")
    label(ws, 14, "EBIT")
    label(ws, 15, "  Margem EBIT", note=True)
    label(ws, 16, "Impostos (EBIT×t)")
    label(ws, 17, "NOPAT = EBIT(1−t)")
    for i, col in enumerate(YCOL):
        r = rows[i]
        put_in(ws, 9, 3 + i, round(r["vn"] / K, 4), fmt=EURMIL, source=("← M6: DR" if i == 0 else None))
        if i == 0:
            put_f(ws, 10, 3, '=""', fmt=PCT1)
        else:
            prev = YCOL[i - 1]
            put_f(ws, 10, 3 + i, f'=IFERROR(({col}9-{prev}9)/{prev}9,"")', fmt=PCT1)
        put_in(ws, 11, 3 + i, round(r["ebitda"] / K, 4), fmt=EURMIL, source=("← M6: DR" if i == 0 else None))
        put_f(ws, 12, 3 + i, f'=IFERROR({col}11/{col}9,"")', fmt=PCT1)
        put_in(ws, 13, 3 + i, round(r["da"] / K, 4), fmt=EURMIL, source=("← M6" if i == 0 else None))
        put_f(ws, 14, 3 + i, f"={col}11-{col}13", fmt=EURMIL)
        put_f(ws, 15, 3 + i, f'=IFERROR({col}14/{col}9,"")', fmt=PCT1)
        put_f(ws, 16, 3 + i, f"=-{col}14*{col}7", fmt=EURMIL)
        put_f(ws, 17, 3 + i, f"={col}14*(1-{col}7)", fmt=EURMIL)

    section(ws, 18, "C. CÁLCULO DO FCFF (€ mil)", span=1 + NY)
    label(ws, 19, "NOPAT")
    label(ws, 20, "(+) Depreciações & Amortizações")
    label(ws, 21, "(−) CapEx")
    label(ws, 22, "(−) Δ Fundo de Maneio (ΔNWC)")
    label(ws, 23, "FCFF", bold=True)
    for i, col in enumerate(YCOL):
        r = rows[i]
        put_f(ws, 19, 3 + i, f"={col}17", fmt=EURMIL)
        put_f(ws, 20, 3 + i, f"={col}13", fmt=EURMIL)
        put_in(ws, 21, 3 + i, round(r["capex"] / K, 4), fmt=EURMIL, source=("← M6: DFC" if i == 0 else None))
        put_in(ws, 22, 3 + i, round(r["dnwc"] / K, 4), fmt=EURMIL, source=("← M6: Δ NFM" if i == 0 else None))
        put_f(ws, 23, 3 + i, f"={col}19+{col}20-{col}21-{col}22", fmt=EURMIL)

    section(ws, 24, "D. DESCONTO E VALOR PRESENTE (€ mil)", span=1 + NY)
    label(ws, 25, "Fator de Desconto")
    label(ws, 26, "FCFF Descontado")
    for i, col in enumerate(YCOL):
        put_f(ws, 25, 3 + i, f"=1/(1+{col}6)^{i+1}", fmt=FACT)
        put_f(ws, 26, 3 + i, f"={col}23*{col}25", fmt=EURMIL)

    section(ws, 27, "E. VALOR TERMINAL E ENTERPRISE VALUE (€ mil)", span=1 + NY)
    label(ws, 28, "FCFF Terminal (n+1) = FCFF₂₀₃₄×(1+gₙ)")
    put_f(ws, 28, 3, f"={LAST}23*(1+'{S_PARAMS}'!C25)", fmt=EURMIL)
    label(ws, 29, "Valor Terminal = FCFF(n+1)/(WACC−gₙ)")
    put_f(ws, 29, 3, f"=IFERROR(C28/('{S_PARAMS}'!C23-'{S_PARAMS}'!C25),\"\")", fmt=EURMIL)
    label(ws, 30, "Fator Desconto VT (ano 10)")
    put_f(ws, 30, 3, f"={LAST}25", fmt=FACT)
    label(ws, 31, "Valor Atual do Valor Terminal")
    put_f(ws, 31, 3, "=C29*C30", fmt=EURMIL)
    label(ws, 32, "Soma FCFF Descontados (anos 1-10)")
    put_f(ws, 32, 3, f"=SUM(C26:{LAST}26)", fmt=EURMIL)
    label(ws, 33, "Enterprise Value (EV)", bold=True)
    put_f(ws, 33, 3, "=C31+C32", fmt=EURMIL)
    label(ws, 34, "(−) Dívida Líquida")
    put_f(ws, 34, 3, f"='{S_PARAMS}'!C36", fmt=EURMIL, ref=True)
    label(ws, 35, "Equity Value", bold=True)
    put_f(ws, 35, 3, "=C33-C34", fmt=EURMIL)
    label(ws, 36, "Nº Ações em Circulação")
    put_f(ws, 36, 3, f"='{S_PARAMS}'!C37", fmt="#,##0", ref=True)
    label(ws, 37, "Valor por Ação (€)", bold=True)
    put_f(ws, 37, 3, '=IFERROR(C35*1000/C36,"")', fmt=EUR2, source="€mil→€: equity×1000/ações")
    ws.cell(37, 4, "€/ação").font = F_NOTE


def build_mult(ws, E):
    ws.column_dimensions["B"].width = 40
    for col in ["C", "D", "E"]: ws.column_dimensions[col].width = 16
    ws.column_dimensions["F"].width = 42
    book_eq = E["book_eq"] / K
    title(ws, "AVALIAÇÃO POR MÚLTIPLOS — AVALIAÇÃO RELATIVA", span=5)
    label(ws, 2, "Métrica forward 2025E ← DCF. Síntese = MÉDIA dos 5 múltiplos (igual ao motor M6).", note=True)
    head(ws, 3, [(2, "Múltiplo / Item"), (3, "Métrica Empresa"), (4, "Setor"),
                 (5, "Múltiplo Aplicado"), (6, "Notas")])
    rows = [
        (4, "EV/EBITDA (2025E)", f"='{S_DCF}'!C11", f"='{S_PARAMS}'!C39", "EBITDA forward 2025E ← DCF"),
        (5, "EV/EBIT (2025E)", f"='{S_DCF}'!C14", f"='{S_PARAMS}'!C40", "EBIT forward 2025E ← DCF"),
        (6, "P/E (2025E)", f"='{S_DCF}'!C17", f"='{S_PARAMS}'!C41", "Lucro normalizado = EBIT×(1−t) ← DCF (coerente c/ motor)"),
        (7, "P/BV", round(book_eq, 1), f"='{S_PARAMS}'!C42", "Capital próprio CONTABILÍSTICO ← M6 total CP 2024"),
        (8, "EV/Sales (2025E)", f"='{S_DCF}'!C9", f"='{S_PARAMS}'!C43", "Receita forward 2025E ← DCF"),
    ]
    for r, nome, metric, med, nota in rows:
        label(ws, r, nome)
        if isinstance(metric, str):
            put_f(ws, r, 3, metric, fmt=EURMIL, ref=True)
        else:
            put_in(ws, r, 3, metric, fmt=EURMIL)
        put_f(ws, r, 4, med, fmt=MULT, ref=True)
        put_f(ws, r, 5, f"=D{r}", fmt=MULT)
        ws.cell(r, 6, nota).font = F_NOTE

    section(ws, 10, "EQUITY VALUE POR MÚLTIPLO (€ mil) — dívida líquida = conceito único", span=5)
    # Equity (col D) homogéneo para a média dos 5
    label(ws, 11, "EV/EBITDA → EV implícito"); put_f(ws, 11, 3, "=C4*E4", fmt=EURMIL)
    put_f(ws, 11, 4, f"=C11-'{S_PARAMS}'!C36", fmt=EURMIL, ref=True); ws.cell(11, 5, "Equity").font = F_NOTE
    label(ws, 12, "EV/EBIT → EV implícito"); put_f(ws, 12, 3, "=C5*E5", fmt=EURMIL)
    put_f(ws, 12, 4, f"=C12-'{S_PARAMS}'!C36", fmt=EURMIL, ref=True); ws.cell(12, 5, "Equity").font = F_NOTE
    label(ws, 13, "P/E → Equity (já é equity)"); put_f(ws, 13, 3, "=C6*E6", fmt=EURMIL)
    put_f(ws, 13, 4, "=C13", fmt=EURMIL); ws.cell(13, 5, "Equity").font = F_NOTE
    label(ws, 14, "P/BV → Equity (já é equity)"); put_f(ws, 14, 3, "=C7*E7", fmt=EURMIL)
    put_f(ws, 14, 4, "=C14", fmt=EURMIL); ws.cell(14, 5, "Equity").font = F_NOTE
    label(ws, 15, "EV/Sales → EV implícito"); put_f(ws, 15, 3, "=C8*E8", fmt=EURMIL)
    put_f(ws, 15, 4, f"=C15-'{S_PARAMS}'!C36", fmt=EURMIL, ref=True); ws.cell(15, 5, "Equity").font = F_NOTE

    label(ws, 17, "Síntese — Equity MÉDIO dos 5 múltiplos", bold=True)
    ws.cell(17, 2).fill = FILL_SECT
    put_f(ws, 17, 4, "=AVERAGE(D11,D12,D13,D14,D15)", fmt=EURMIL)
    ws.cell(17, 5, "→ Equity p/ Síntese").font = F_NOTE


def build_fcfe(ws, E):
    ws.column_dimensions["B"].width = 42
    for col in YCOL: ws.column_dimensions[col].width = 11
    rows = E["rows"]
    title(ws, "AVALIAÇÃO POR FCFE — PONTE FCFF → EQUITY · 10 anos", span=1 + NY)
    label(ws, 2, "FCFE = FCFF − Juros×(1−t) + ΔDívida Financeira (ponte Damodaran). Séries ← M6.", note=True)
    head(ws, 4, [(2, "Item")] + [(3 + i, f"Ano {i+1}\n{y}") for i, y in enumerate(YEARS)])

    section(ws, 5, "A. PRESSUPOSTOS", span=1 + NY)
    label(ws, 6, "ke (desconto)")
    for i, col in enumerate(YCOL):
        put_f(ws, 6, 3 + i, f"='{S_PARAMS}'!$C$15", fmt=PCT, ref=True)

    section(ws, 8, "B. PROJEÇÃO FCFE (€ mil)", span=1 + NY)
    label(ws, 9, "FCFF (← DCF-FCFF)")
    label(ws, 10, "Juros Financeiros (kd×Dívida)")
    label(ws, 11, "(−) Juros após imposto [Juros×(1−t)]")
    label(ws, 12, "(+) Variação de Dívida Financeira")
    label(ws, 13, "FCFE", bold=True)
    for i, col in enumerate(YCOL):
        r = rows[i]
        put_f(ws, 9, 3 + i, f"='{S_DCF}'!{col}23", fmt=EURMIL, ref=True)
        put_in(ws, 10, 3 + i, round(r["juros"] / K, 4), fmt=EURMIL, source=("← M6: DR juros" if i == 0 else None))
        put_f(ws, 11, 3 + i, f"={col}10*(1-'{S_PARAMS}'!$C$16)", fmt=EURMIL)
        put_in(ws, 12, 3 + i, round(r["net_borrow"] / K, 4), fmt=EURMIL, source=("← M6: Δ dívida fin." if i == 0 else None))
        put_f(ws, 13, 3 + i, f"={col}9-{col}11+{col}12", fmt=EURMIL)

    section(ws, 14, "C. DESCONTO E EQUITY VALUE (€ mil)", span=1 + NY)
    label(ws, 15, "Fator de Desconto (ke)")
    label(ws, 16, "FCFE Descontado")
    for i, col in enumerate(YCOL):
        put_f(ws, 15, 3 + i, f"=1/(1+{col}6)^{i+1}", fmt=FACT)
        put_f(ws, 16, 3 + i, f"={col}13*{col}15", fmt=EURMIL)
    label(ws, 17, "FCFE Terminal (n+1) = FCFE₂₀₃₄×(1+gₙ)")
    put_f(ws, 17, 3, f"={LAST}13*(1+'{S_PARAMS}'!C25)", fmt=EURMIL)
    label(ws, 18, "Valor Terminal FCFE = (n+1)/(ke−gₙ)")
    put_f(ws, 18, 3, f"=IFERROR(C17/('{S_PARAMS}'!C15-'{S_PARAMS}'!C25),\"\")", fmt=EURMIL)
    label(ws, 19, "Valor Atual do Valor Terminal")
    put_f(ws, 19, 3, f"=C18*{LAST}15", fmt=EURMIL)
    label(ws, 20, "Soma FCFE Descontados (1-10)")
    put_f(ws, 20, 3, f"=SUM(C16:{LAST}16)", fmt=EURMIL)
    label(ws, 21, "Equity Value (FCFE)", bold=True)
    put_f(ws, 21, 3, "=C19+C20", fmt=EURMIL)
    label(ws, 22, "Valor por Ação (€)", bold=True)
    put_f(ws, 22, 3, f"=IFERROR(C21*1000/'{S_PARAMS}'!C37,\"\")", fmt=EUR2)
    ws.cell(22, 4, "€/ação").font = F_NOTE


def build_stress(ws):
    ws.column_dimensions["B"].width = 18
    for col in YCOL[:5]: ws.column_dimensions[col].width = 14
    title(ws, "ANÁLISE DE SENSIBILIDADE — STRESS TESTS (dinâmico, 10 anos)", span=6)
    label(ws, 3, "Equity Value DCF-FCFF em função de WACC e gₙ (Δ p.p.) — fórmula viva por célula.")
    g_lbl = ["-1.0%", "-0.5%", "+0.0%", "+0.5%", "+1.0%"]
    w_lbl = ["-2.0%", "-1.0%", "+0.0%", "+1.0%", "+2.0%"]
    g_off = [-0.01, -0.005, 0.0, 0.005, 0.01]
    w_off = [-0.02, -0.01, 0.0, 0.01, 0.02]
    head(ws, 4, [(2, "Δ WACC \\ Δgₙ →")] + [(3 + j, g_lbl[j]) for j in range(5)])
    for i in range(5):
        c = ws.cell(5 + i, 2, w_lbl[i])
        c.font = F_HEAD; c.fill = FILL_HEAD; c.alignment = ALIGN_C; c.border = BORDER

    base_w = f"'{S_PARAMS}'!$C$23"
    base_g = f"'{S_PARAMS}'!$C$25"
    nd = f"'{S_PARAMS}'!$C$36"
    fcff = [f"'{S_DCF}'!${c}$23" for c in YCOL]   # FCFF anos 1-10
    for i, w in enumerate(w_off):
        for j, g in enumerate(g_off):
            W = f"({base_w}+({w}))"
            G = f"({base_g}+({g}))"
            pv = "+".join(f"{fcff[t]}/(1+{W})^{t+1}" for t in range(NY))
            tv = f"({fcff[-1]}*(1+{G})/({W}-{G}))/(1+{W})^{NY}"
            cell = ws.cell(5 + i, 3 + j, f'=IFERROR(({pv}+{tv})-{nd},"")')
            cell.font = F_CALC; cell.number_format = EURMIL
            cell.alignment = ALIGN_R; cell.border = BORDER
    label(ws, 11, "Cada célula recalcula o Equity Value (DCF) ao WACC e gₙ indicados — "
                  "totalmente dinâmico, sem valores fixos.", note=True)


def build_syn(ws):
    ws.column_dimensions["B"].width = 46
    for col in ["C", "D", "E"]: ws.column_dimensions[col].width = 18
    title(ws, "SÍNTESE — COMPARAÇÃO E PREÇO DE NEGOCIAÇÃO", span=4)
    label(ws, 2, "Hierarquia de fiabilidade: intrínseco pesa o preço (DCF-FCFF 60 % + "
                 "FCFE 40 %); múltiplos = banda de confronto (0 %). Desconto −10 %.", note=True)
    head(ws, 3, [(2, "Metodologia"), (3, "Equity Value (€ mil)"), (4, "Peso (%)"),
                 (5, "Contribuição (€ mil)")])
    label(ws, 4, "1. DCF — FCFF (intrínseco · ótica da empresa)")
    put_f(ws, 4, 3, f"='{S_DCF}'!C35", fmt=EURMIL, ref=True)
    put_in(ws, 4, 4, 0.60, fmt=PCT, source="intrínseco — isola valor do Hub da estrutura de capital")
    put_f(ws, 4, 5, "=C4*D4", fmt=EURMIL)
    label(ws, 5, "2. FCFE (intrínseco · ótica do acionista)")
    put_f(ws, 5, 3, f"='{S_FCFE}'!C21", fmt=EURMIL, ref=True)
    put_in(ws, 5, 4, 0.40, fmt=PCT, source="intrínseco — capta dividendos pós-serviço da dívida")
    put_f(ws, 5, 5, "=C5*D5", fmt=EURMIL)
    label(ws, 6, "3. Múltiplos (média de 5) — banda de confronto")
    put_f(ws, 6, 3, f"='{S_MULT}'!D17", fmt=EURMIL, ref=True)
    put_in(ws, 6, 4, 0.00, fmt=PCT, source="relativo — sanity-check, não contamina o intrínseco (Grestel privada)")
    put_f(ws, 6, 5, "=C6*D6", fmt=EURMIL)
    label(ws, 7, "VALOR PONDERADO (só intrínseco)", bold=True)
    put_f(ws, 7, 4, "=SUM(D4:D6)", fmt=PCT)
    put_f(ws, 7, 5, '=IFERROR(SUMPRODUCT(C4:C6,D4:D6)/SUM(D4:D6),"")', fmt=EURMIL)
    label(ws, 8, "Confronto: múltiplos / valor intrínseco (sanity-check)", note=True)
    put_f(ws, 8, 5, '=IFERROR(C6/E7,"")', fmt=PCT)
    ws.cell(8, 4, "→ próximo de 100% valida o intrínseco").font = F_NOTE

    section(ws, 9, "DECISÃO FINAL — PREÇO MÍNIMO DE NEGOCIAÇÃO", span=4)
    label(ws, 10, "Valor Ponderado de Referência (€ mil)"); put_f(ws, 10, 3, "=E7", fmt=EURMIL)
    label(ws, 11, "Desconto de Negociação (%)")
    put_in(ws, 11, 3, -0.10, fmt=PCT, source="margem negocial (motor M6: −10%)")
    label(ws, 12, "Preço Mínimo (€ mil)", bold=True); put_f(ws, 12, 3, "=C10*(1+C11)", fmt=EURMIL)
    label(ws, 13, "Preço Mínimo por Ação (€)", bold=True)
    put_f(ws, 13, 3, f"=IFERROR(C12*1000/'{S_PARAMS}'!C37,\"\")", fmt=EUR2)
    ws.cell(13, 4, "€/ação").font = F_NOTE


def build(out_path, cenario="Base"):
    E = extract_engine(cenario)
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    build_idx(wb.create_sheet(S_IDX))
    build_params(wb.create_sheet(S_PARAMS), E)
    build_dcf(wb.create_sheet(S_DCF), E)
    build_mult(wb.create_sheet(S_MULT), E)
    build_fcfe(wb.create_sheet(S_FCFE), E)
    build_stress(wb.create_sheet(S_STRESS))
    build_syn(wb.create_sheet(S_SYN))
    try:
        from openpyxl.workbook.properties import CalcProperties
        wb.calculation = CalcProperties(fullCalcOnLoad=True)
    except Exception:
        pass
    wb.save(out_path)
    return E


if __name__ == "__main__":
    import sys
    out = sys.argv[1] if len(sys.argv) > 1 else os.path.join(
        os.path.dirname(__file__), "..", "oe5_markdowns", "OE05_G18_260601_PREENCHIDO.xlsx")
    out = os.path.normpath(out)
    E = build(out)
    print("Gerado:", out)
