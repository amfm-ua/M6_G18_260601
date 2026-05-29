"""Gerador do template estrutural da avaliação OE5 (sem números).

Cria oe5_markdowns/OE05_G18_260601_estrutura.xlsx com:
  - todas as células de INPUT vazias (azul) — a preencher pelo grupo ou pelo motor M6
  - todas as FÓRMULAS ligadas (modelo dinâmico e interligado)
  - endereços de célula compatíveis com src/engine/valuation/excel_reader.py
  - WACC com benefício fiscal da dívida (obrigatório OE5)
  - grelha de stress WACC x g totalmente dinâmica (fórmula viva por célula)

Sem dados: serve para o utilizador resolver o exercício.  Os 5 bugs da versão
anterior ficam estruturalmente impossíveis (ver README do projeto / memória OE5).
"""
from __future__ import annotations

import os

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ── Paleta e estilos ──────────────────────────────────────────────────────────
C_INPUT = "1F4FD0"     # azul   — input manual (a preencher)
C_REF = "1F8A4C"       # verde  — referência a outra folha
C_HEAD_BG = "1F3A5F"   # fundo cabeçalho
C_SECT_BG = "DCE6F1"   # fundo secção
C_TITLE_BG = "2E5496"

F_TITLE = Font(name="Calibri", size=13, bold=True, color="FFFFFF")
F_HEAD = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
F_SECT = Font(name="Calibri", size=10, bold=True, color="1F3A5F")
F_INPUT = Font(name="Calibri", size=10, color=C_INPUT, bold=True)
F_REF = Font(name="Calibri", size=10, color=C_REF)
F_CALC = Font(name="Calibri", size=10, color="000000")
F_LABEL = Font(name="Calibri", size=10, color="000000")
F_NOTE = Font(name="Calibri", size=8, italic=True, color="666666")

FILL_HEAD = PatternFill("solid", fgColor=C_HEAD_BG)
FILL_SECT = PatternFill("solid", fgColor=C_SECT_BG)
FILL_TITLE = PatternFill("solid", fgColor=C_TITLE_BG)

_thin = Side(style="thin", color="BFBFBF")
BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)

PCT = "0.00%"
PCT1 = "0.0%"
EURMIL = "#,##0"
EUR2 = "#,##0.00"
MULT = '0.00"x"'
FACT = "0.0000"

# Formatos numéricos por chave de tipo
ALIGN_L = Alignment(horizontal="left", vertical="center", wrap_text=False)
ALIGN_R = Alignment(horizontal="right", vertical="center")
ALIGN_C = Alignment(horizontal="center", vertical="center", wrap_text=True)


def title(ws, text, span=6):
    ws.merge_cells(start_row=1, start_column=2, end_row=1, end_column=1 + span)
    c = ws.cell(1, 2, text)
    c.font = F_TITLE
    c.fill = FILL_TITLE
    c.alignment = ALIGN_L
    ws.row_dimensions[1].height = 22


def label(ws, r, text, *, col=2, bold=False, note=False):
    c = ws.cell(r, col, text)
    c.font = F_NOTE if note else (F_SECT if bold else F_LABEL)
    c.alignment = ALIGN_L
    return c


def section(ws, r, text, span=6):
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=1 + span)
    c = ws.cell(r, 2, text)
    c.font = F_SECT
    c.fill = FILL_SECT
    c.alignment = ALIGN_L


def head(ws, r, cols):
    for col, text in cols:
        c = ws.cell(r, col, text)
        c.font = F_HEAD
        c.fill = FILL_HEAD
        c.alignment = ALIGN_C
        c.border = BORDER


def put_input(ws, r, c, *, fmt=None, unit=None, source=None):
    cell = ws.cell(r, c)              # VAZIO — input a preencher
    cell.font = F_INPUT
    cell.alignment = ALIGN_R
    cell.border = BORDER
    if fmt:
        cell.number_format = fmt
    if unit is not None:
        ws.cell(r, c + 1, unit).font = F_NOTE
    if source is not None:
        ws.cell(r, c + 2, source).font = F_NOTE
    return cell


def put_formula(ws, r, c, formula, *, fmt=None, ref=False, unit=None, source=None):
    cell = ws.cell(r, c, formula)
    cell.font = F_REF if ref else F_CALC
    cell.alignment = ALIGN_R
    cell.border = BORDER
    if fmt:
        cell.number_format = fmt
    if unit is not None:
        ws.cell(r, c + 1, unit).font = F_NOTE
    if source is not None:
        ws.cell(r, c + 2, source).font = F_NOTE
    return cell


# Nomes EXATOS exigidos por excel_reader.py
S_IDX = "📋 Índice"
S_PARAMS = "⚙️ Pressupostos"
S_DCF = "📊 DCF-FCFF"
S_MULT = "📈 Múltiplos"
S_FCFE = "💰 FCFE"
S_STRESS = "📉 Stress Tests"
S_SYN = "🏆 Síntese"

YEAR_COLS = ["C", "D", "E", "F", "G"]
YEARS = [2025, 2026, 2027, 2028, 2029]


def build():
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    _idx(wb.create_sheet(S_IDX))
    _params(wb.create_sheet(S_PARAMS))
    _dcf(wb.create_sheet(S_DCF))
    _mult(wb.create_sheet(S_MULT))
    _fcfe(wb.create_sheet(S_FCFE))
    _stress(wb.create_sheet(S_STRESS))
    _syn(wb.create_sheet(S_SYN))

    out = os.path.join(
        os.path.dirname(__file__), "..", "oe5_markdowns",
        "OE05_G18_260601_estrutura.xlsx",
    )
    out = os.path.normpath(out)
    wb.save(out)
    return out


# ── Índice ─────────────────────────────────────────────────────────────────────
def _idx(ws):
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 60
    ws.column_dimensions["D"].width = 16
    title(ws, "OE05 — MODELO DE AVALIAÇÃO · ESTRUTURA (sem dados)", span=3)
    label(ws, 2, "Grestel — Produtos Cerâmicos, S.A.   ·   PEF 2025-26   ·   Grupo 18")
    head(ws, 4, [(2, "Folha"), (3, "Conteúdo"), (4, "Estado")])
    rows = [
        (S_IDX, "Navegação e legenda", "—"),
        (S_PARAMS, "Inputs centrais: WACC (c/ benefício fiscal), taxas, beta, base, múltiplos", "🟡 Preencher"),
        (S_DCF, "DCF-FCFF (obrigatório) — projeção ← M6, Enterprise/Equity Value", "🟡 Preencher"),
        (S_MULT, "Avaliação relativa — EV/EBITDA, EV/EBIT, EV/Sales, P/E, P/BV", "🔄 Ligado"),
        (S_FCFE, "Ótica do acionista — FCFE descontado a ke", "🟡 Preencher"),
        (S_STRESS, "Sensibilidade dinâmica WACC × g (fórmula viva)", "🔄 Automático"),
        (S_SYN, "Comparação dos 3 métodos + preço mínimo de negociação", "🔄 Automático"),
    ]
    r = 5
    for nome, cont, estado in rows:
        label(ws, r, nome)
        ws.cell(r, 3, cont).font = F_LABEL
        ws.cell(r, 4, estado).alignment = ALIGN_C
        for cc in (2, 3, 4):
            ws.cell(r, cc).border = BORDER
        r += 1
    r += 1
    section(ws, r, "LEGENDA", span=3); r += 1
    for txt, fnt in [
        ("● Azul — input manual a preencher (ou alimentado pelo motor M6)", F_INPUT),
        ("● Verde — referência a outra folha", F_REF),
        ("● Preto — fórmula calculada", F_CALC),
    ]:
        c = ws.cell(r, 2, txt); c.font = fnt; r += 1
    r += 1
    label(ws, r, "Nota: ficheiro estrutural sem números. Endereços compatíveis com "
                 "src/engine/valuation/excel_reader.py (o M6 pode preencher).", note=True)


# ── Pressupostos ────────────────────────────────────────────────────────────────
def _params(ws):
    for col, w in {"B": 42, "C": 16, "D": 10, "E": 60}.items():
        ws.column_dimensions[col].width = w
    title(ws, "⚙️  PRESSUPOSTOS E PARÂMETROS DO MODELO", span=4)
    label(ws, 2, "Inputs a AZUL (vazios). Fórmulas calculam automaticamente.", note=True)
    head(ws, 3, [(2, "Parâmetro"), (3, "Valor"), (4, "Unid."), (5, "Fonte / Notas")])

    section(ws, 4, "A. EMPRESA")
    label(ws, 5, "Nome da Empresa")
    ws.cell(5, 3, "Grestel — Produtos Cerâmicos, S.A.").font = F_INPUT
    ws.cell(5, 5, "Identificação").font = F_NOTE
    label(ws, 6, "Setor / Indústria")
    ws.cell(6, 5, "← M6 / Damodaran: Household Products (Europe)").font = F_NOTE
    label(ws, 7, "Moeda de Reporte"); ws.cell(7, 4, "EUR").font = F_NOTE
    label(ws, 8, "Ano Base (histórico)")
    put_input(ws, 8, 3, unit="Ano", source="Último ano real (R&C) ← M6")
    label(ws, 9, "Primeiro Ano de Projeção")
    put_formula(ws, 9, 3, "=C8+1", unit="Ano")
    label(ws, 10, "Horizonte de Projeção (anos)")
    put_input(ws, 10, 3, unit="Anos", source="Período explícito do DCF")

    section(ws, 11, "B. CUSTO DE CAPITAL (WACC)")
    label(ws, 12, "Taxa Livre de Risco (Rf)")
    put_input(ws, 12, 3, fmt=PCT, unit="%", source="← Mercado: OT PT 10a (BPstat)")
    label(ws, 13, "Prémio de Risco de Mercado (ERP)")
    put_input(ws, 13, 3, fmt=PCT, unit="%", source="← Mercado: Damodaran ctry premium PT")
    label(ws, 14, "Beta Alavancado (β)")
    put_input(ws, 14, 3, fmt="0.00", unit="—", source="← Mercado: βU setor re-alavancado c/ D/E Grestel")
    label(ws, 15, "Custo Capital Próprio (ke)")
    put_formula(ws, 15, 3, "=C12+C14*C13", fmt=PCT, unit="%", source="CAPM: ke = Rf + β×ERP")
    label(ws, 16, "Taxa de Imposto (t)")
    put_input(ws, 16, 3, fmt=PCT, unit="%", source="IRC vigente ← M6")
    label(ws, 17, "Custo Dívida Bruta (kd)")
    put_input(ws, 17, 3, fmt=PCT, unit="%", source="← M6: taxa média da dívida")
    label(ws, 18, "Custo Dívida Após Impostos")
    put_formula(ws, 18, 3, "=C17*(1-C16)", fmt=PCT, unit="%", source="kd×(1−t)  ← BENEFÍCIO FISCAL")
    label(ws, 19, "Dívida (valor mercado)")
    put_input(ws, 19, 3, fmt=EURMIL, unit="€ mil", source="← M6: empréstimos fin. (pesos WACC)")
    label(ws, 20, "Capital Próprio (valor mercado)")
    put_input(ws, 20, 3, fmt=EURMIL, unit="€ mil", source="← M6: capital próprio")
    label(ws, 21, "Peso Dívida (D/V)")
    put_formula(ws, 21, 3, '=IFERROR(C19/(C19+C20),"")', fmt=PCT, unit="%")
    label(ws, 22, "Peso Capital Próprio (E/V)")
    put_formula(ws, 22, 3, '=IFERROR(C20/(C19+C20),"")', fmt=PCT, unit="%")
    label(ws, 23, "WACC", bold=True)
    put_formula(ws, 23, 3, "=C15*C22+C18*C21", fmt=PCT, unit="%",
                source="WACC = ke×E/V + kd×(1−t)×D/V")

    section(ws, 24, "C. CRESCIMENTO E VALOR TERMINAL")
    label(ws, 25, "Crescimento Fase 1 (g1) — anos 1-3")
    put_input(ws, 25, 3, fmt=PCT, unit="%", source="← M6: crescimento de receita")
    label(ws, 26, "Crescimento Fase 2 (g2) — anos 4-5")
    put_input(ws, 26, 3, fmt=PCT, unit="%", source="← M6: crescimento de receita")
    label(ws, 27, "Crescimento Terminal (gn)")
    put_input(ws, 27, 3, fmt=PCT, unit="%", source="≤ crescimento nominal da economia (Damodaran cap.12)")
    label(ws, 28, "Taxa Reinvestimento Terminal")
    put_formula(ws, 28, 3, '=IFERROR(C27/C29,"")', fmt=PCT, unit="%", source="= gn / ROC (coerência Damodaran)")
    label(ws, 29, "Retorno sobre Capital (ROC)")
    put_input(ws, 29, 3, fmt=PCT, unit="%", source="← M6: ROIC histórico / setor")

    section(ws, 30, "D. DADOS FINANCEIROS BASE (€ mil)")
    label(ws, 31, "EBIT"); put_input(ws, 31, 3, fmt=EURMIL, unit="€ mil", source="← M6: DR ano base")
    label(ws, 32, "Depreciações & Amortizações"); put_input(ws, 32, 3, fmt=EURMIL, unit="€ mil", source="← M6: mapa D&A")
    label(ws, 33, "CapEx"); put_input(ws, 33, 3, fmt=EURMIL, unit="€ mil", source="← M6: mapa de investimentos")
    label(ws, 34, "Variação Fundo de Maneio (ΔNWC)"); put_input(ws, 34, 3, fmt=EURMIL, unit="€ mil", source="← M6: Δ NFM")
    label(ws, 35, "EBITDA"); put_formula(ws, 35, 3, "=C31+C32", fmt=EURMIL, unit="€ mil", source="= EBIT + D&A")
    label(ws, 36, "Dívida Líquida", bold=True)
    put_input(ws, 36, 3, fmt=EURMIL, unit="€ mil", source="← M6: dívida fin. − caixa (CONCEITO ÚNICO)")
    label(ws, 37, "Nº de Ações em Circulação", bold=True)
    put_input(ws, 37, 3, fmt="#,##0", unit="unid.", source="← R&C: capital social / valor nominal")

    section(ws, 38, "E. MÚLTIPLOS DE REFERÊNCIA (SETOR)")
    for r, nome in [
        (39, "EV/EBITDA — Mediana Setor"),
        (40, "EV/EBIT — Mediana Setor"),
        (41, "P/E — Mediana Setor"),
        (42, "P/BV — Mediana Setor"),
        (43, "EV/Sales — Mediana Setor"),
    ]:
        label(ws, r, nome)
        put_input(ws, r, 3, fmt=MULT, unit="x", source="← Mercado: Damodaran (setor europeu)")


# ── DCF-FCFF ────────────────────────────────────────────────────────────────────
def _dcf(ws):
    ws.column_dimensions["B"].width = 40
    for col in YEAR_COLS + ["H"]:
        ws.column_dimensions[col].width = 12
    title(ws, "📊  MODELO DCF — FREE CASH FLOW TO THE FIRM (FCFF)", span=7)
    label(ws, 2, "Séries de projeção (azul) ← motor M6. Valuation = fórmulas.", note=True)
    head(ws, 3, [(2, "Item")] + [(3 + i, f"Ano {i+1}\n{y}") for i, y in enumerate(YEARS)] + [(8, "Terminal")])

    section(ws, 4, "A. PRESSUPOSTOS (ligados a ⚙️ Pressupostos)", span=7)
    label(ws, 5, "Taxa de Crescimento Receita")
    for i, col in enumerate(YEAR_COLS):
        put_input(ws, 5, 3 + i, fmt=PCT, source=("← M6" if i == 0 else None))
    put_formula(ws, 5, 8, f"='{S_PARAMS}'!C27", fmt=PCT, ref=True)
    label(ws, 6, "WACC (desconto)")
    for i, col in enumerate(YEAR_COLS):
        put_formula(ws, 6, 3 + i, f"='{S_PARAMS}'!$C$23", fmt=PCT, ref=True)
    put_formula(ws, 6, 8, f"='{S_PARAMS}'!$C$23", fmt=PCT, ref=True)
    label(ws, 7, "Taxa de Imposto (t)")
    for i, col in enumerate(YEAR_COLS):
        put_formula(ws, 7, 3 + i, f"='{S_PARAMS}'!$C$16", fmt=PCT, ref=True)
    put_formula(ws, 7, 8, f"='{S_PARAMS}'!$C$16", fmt=PCT, ref=True)

    section(ws, 8, "B. PROJEÇÃO DA DEMONSTRAÇÃO DE RESULTADOS (€ mil)", span=7)
    label(ws, 9, "Receita (Volume de Negócios)")
    put_input(ws, 9, 2, source=None)            # B9 = receita ano base (input)
    ws.cell(9, 2).value = None
    label(ws, 9, "Receita (Volume de Negócios)")
    for i in range(5):
        put_input(ws, 9, 3 + i, fmt=EURMIL, source=("← M6: DR" if i == 0 else None))
    put_formula(ws, 9, 8, "=G9*(1+H5)", fmt=EURMIL)
    label(ws, 10, "  Crescimento YoY", note=True)
    put_formula(ws, 10, 3, '=IFERROR((C9-B9)/B9,"")', fmt=PCT1)
    for i in range(1, 5):
        cur, prev = YEAR_COLS[i], YEAR_COLS[i - 1]
        put_formula(ws, 10, 3 + i, f'=IFERROR(({cur}9-{prev}9)/{prev}9,"")', fmt=PCT1)
    label(ws, 11, "EBITDA")
    for i in range(5):
        put_input(ws, 11, 3 + i, fmt=EURMIL, source=("← M6: DR" if i == 0 else None))
    put_formula(ws, 11, 8, "=G11*(1+H5)", fmt=EURMIL)
    label(ws, 12, "  Margem EBITDA", note=True)
    for i, col in enumerate(YEAR_COLS):
        put_formula(ws, 12, 3 + i, f'=IFERROR({col}11/{col}9,"")', fmt=PCT1)
    label(ws, 13, "Depreciações (D&A)")
    for i in range(5):
        put_input(ws, 13, 3 + i, fmt=EURMIL, source=("← M6" if i == 0 else None))
    put_formula(ws, 13, 8, "=G13", fmt=EURMIL)
    label(ws, 14, "EBIT")
    for i, col in enumerate(YEAR_COLS):
        put_formula(ws, 14, 3 + i, f"={col}11-{col}13", fmt=EURMIL)
    put_formula(ws, 14, 8, "=H11-H13", fmt=EURMIL)
    label(ws, 15, "  Margem EBIT", note=True)
    for i, col in enumerate(YEAR_COLS):
        put_formula(ws, 15, 3 + i, f'=IFERROR({col}14/{col}9,"")', fmt=PCT1)
    label(ws, 16, "Impostos (EBIT×t)")
    for i, col in enumerate(YEAR_COLS):
        put_formula(ws, 16, 3 + i, f"=-{col}14*{col}7", fmt=EURMIL)
    put_formula(ws, 16, 8, "=H14*H7", fmt=EURMIL)
    label(ws, 17, "NOPAT = EBIT(1−t)")
    for i, col in enumerate(YEAR_COLS):
        put_formula(ws, 17, 3 + i, f"={col}14*(1-{col}7)", fmt=EURMIL)
    put_formula(ws, 17, 8, "=H14-H16", fmt=EURMIL)

    section(ws, 18, "C. CÁLCULO DO FCFF (€ mil)", span=7)
    label(ws, 19, "NOPAT")
    for i, col in enumerate(YEAR_COLS):
        put_formula(ws, 19, 3 + i, f"={col}17", fmt=EURMIL)
    label(ws, 20, "(+) Depreciações & Amortizações")
    for i, col in enumerate(YEAR_COLS):
        put_formula(ws, 20, 3 + i, f"={col}13", fmt=EURMIL)
    label(ws, 21, "(−) CapEx")
    for i in range(5):
        put_input(ws, 21, 3 + i, fmt=EURMIL, source=("← M6: DFC" if i == 0 else None))
    label(ws, 22, "(−) Δ Fundo de Maneio (ΔNWC)")
    for i in range(5):
        put_input(ws, 22, 3 + i, fmt=EURMIL, source=("← M6: Δ NFM" if i == 0 else None))
    label(ws, 23, "FCFF", bold=True)
    for i, col in enumerate(YEAR_COLS):
        put_formula(ws, 23, 3 + i, f"={col}19+{col}20-{col}21-{col}22", fmt=EURMIL)

    section(ws, 24, "D. DESCONTO E VALOR PRESENTE (€ mil)", span=7)
    label(ws, 25, "Fator de Desconto")
    for i, col in enumerate(YEAR_COLS):
        put_formula(ws, 25, 3 + i, f"=1/(1+{col}6)^{i+1}", fmt=FACT)
    label(ws, 26, "FCFF Descontado")
    for i, col in enumerate(YEAR_COLS):
        put_formula(ws, 26, 3 + i, f"={col}23*{col}25", fmt=EURMIL)

    section(ws, 27, "E. VALOR TERMINAL E ENTERPRISE VALUE (€ mil)", span=7)
    label(ws, 28, "FCFF Terminal (n+1) — consistente c/ reinvestimento")
    # NOPAT terminal × (1 − taxa reinvestimento terminal)
    put_formula(ws, 28, 3, f"=H17*(1-'{S_PARAMS}'!C28)", fmt=EURMIL)
    label(ws, 29, "Valor Terminal = FCFF(n+1)/(WACC−gn)")
    put_formula(ws, 29, 3, f"=IFERROR(C28/('{S_PARAMS}'!C23-'{S_PARAMS}'!C27),\"\")", fmt=EURMIL)
    label(ws, 30, "Fator Desconto VT (ano n)")
    put_formula(ws, 30, 3, "=G25", fmt=FACT)
    label(ws, 31, "Valor Atual do Valor Terminal")
    put_formula(ws, 31, 3, "=C29*C30", fmt=EURMIL)
    label(ws, 32, "Soma FCFF Descontados (anos 1-n)")
    put_formula(ws, 32, 3, "=SUM(C26:G26)", fmt=EURMIL)
    label(ws, 33, "Enterprise Value (EV)", bold=True)
    put_formula(ws, 33, 3, "=C31+C32", fmt=EURMIL)
    label(ws, 34, "(−) Dívida Líquida")
    put_formula(ws, 34, 3, f"='{S_PARAMS}'!C36", fmt=EURMIL, ref=True)
    label(ws, 35, "Equity Value", bold=True)
    put_formula(ws, 35, 3, "=C33-C34", fmt=EURMIL)
    label(ws, 36, "Nº Ações em Circulação")
    put_formula(ws, 36, 3, f"='{S_PARAMS}'!C37", fmt="#,##0", ref=True)
    label(ws, 37, "Valor por Ação (€)", bold=True)
    # CONVERSÃO €mil → €  : equity(€mil) × 1000 / nº ações
    put_formula(ws, 37, 3, '=IFERROR(C35*1000/C36,"")', fmt=EUR2, source="€mil→€: equity×1000/ações")
    ws.cell(37, 4, "€/ação").font = F_NOTE


# ── Múltiplos ───────────────────────────────────────────────────────────────────
def _mult(ws):
    ws.column_dimensions["B"].width = 38
    for col in ["C", "D", "E"]:
        ws.column_dimensions[col].width = 16
    ws.column_dimensions["F"].width = 40
    title(ws, "📈  AVALIAÇÃO POR MÚLTIPLOS — AVALIAÇÃO RELATIVA", span=5)
    head(ws, 3, [(2, "Múltiplo / Item"), (3, "Métrica Empresa"), (4, "Mediana Setor"),
                 (5, "Valor Aplicado"), (6, "Notas")])
    # Métrica empresa (forward 2025E) ← DCF; mediana ← Pressupostos
    rows = [
        (4, "EV/EBITDA (2025E)", f"='{S_DCF}'!C11", f"='{S_PARAMS}'!C39", "EBITDA forward 2025E ← DCF"),
        (5, "EV/EBIT (2025E)", f"='{S_DCF}'!C14", f"='{S_PARAMS}'!C40", "EBIT forward 2025E ← DCF"),
        (6, "P/E (2025E)", f"='{S_FCFE}'!C9", f"='{S_PARAMS}'!C41", "Lucro líquido 2025E ← FCFE"),
        (7, "P/BV", f"='{S_PARAMS}'!C20", f"='{S_PARAMS}'!C42", "Capital próprio ← Pressupostos"),
        (8, "EV/Sales (2025E)", f"='{S_DCF}'!C9", f"='{S_PARAMS}'!C43", "Receita forward 2025E ← DCF"),
    ]
    for r, nome, metric_f, med_f, nota in rows:
        label(ws, r, nome)
        put_formula(ws, r, 3, metric_f, fmt=EURMIL, ref=True)
        put_formula(ws, r, 4, med_f, fmt=MULT, ref=True)
        put_formula(ws, r, 5, f"=D{r}", fmt=MULT)
        ws.cell(r, 6, nota).font = F_NOTE

    section(ws, 10, "EQUITY VALUE POR MÚLTIPLO (€ mil) — dívida líquida = conceito único", span=5)
    # EV implícito = métrica × múltiplo ; Equity = EV − dívida líquida (Pressupostos C36)
    label(ws, 11, "EV implícito — EV/EBITDA")
    put_formula(ws, 11, 3, "=C4*E4", fmt=EURMIL)
    put_formula(ws, 11, 4, f"=C11-'{S_PARAMS}'!C36", fmt=EURMIL, ref=True)
    ws.cell(11, 5, "Equity Value").font = F_NOTE
    label(ws, 12, "EV implícito — EV/EBIT")
    put_formula(ws, 12, 3, "=C5*E5", fmt=EURMIL)
    put_formula(ws, 12, 4, f"=C12-'{S_PARAMS}'!C36", fmt=EURMIL, ref=True)
    ws.cell(12, 5, "Equity Value").font = F_NOTE
    label(ws, 13, "Equity — P/E (já é equity)")
    put_formula(ws, 13, 3, "=C6*E6", fmt=EURMIL)
    ws.cell(13, 4, "—").alignment = ALIGN_R
    ws.cell(13, 5, "Equity Value").font = F_NOTE
    label(ws, 14, "Equity — P/BV (já é equity)")
    put_formula(ws, 14, 3, "=C7*E7", fmt=EURMIL)
    ws.cell(14, 4, "—").alignment = ALIGN_R
    ws.cell(14, 5, "Equity Value").font = F_NOTE
    label(ws, 15, "EV implícito — EV/Sales")
    put_formula(ws, 15, 3, "=C8*E8", fmt=EURMIL)
    put_formula(ws, 15, 4, f"=C15-'{S_PARAMS}'!C36", fmt=EURMIL, ref=True)
    ws.cell(15, 5, "Equity Value").font = F_NOTE

    label(ws, 17, "Síntese — Equity mediano (EV/EBITDA, EV/EBIT, EV/Sales)", bold=True)
    ws.cell(17, 2).fill = FILL_SECT
    # mediana dos EQUITY values dos 3 múltiplos de EV (rótulo correto)
    put_formula(ws, 17, 4, "=MEDIAN(D11,D12,D15)", fmt=EURMIL)
    ws.cell(17, 5, "→ Equity p/ Síntese").font = F_NOTE


# ── FCFE ────────────────────────────────────────────────────────────────────────
def _fcfe(ws):
    ws.column_dimensions["B"].width = 40
    for col in YEAR_COLS + ["H"]:
        ws.column_dimensions[col].width = 12
    title(ws, "💰  AVALIAÇÃO POR FCFE (FREE CASH FLOW TO EQUITY)", span=7)
    label(ws, 2, "FCFE = NI − (CapEx−D&A) − ΔNWC + (Nova Dívida − Amortização). Séries ← M6.", note=True)
    head(ws, 4, [(2, "Item")] + [(3 + i, f"Ano {i+1}\n{y}") for i, y in enumerate(YEARS)] + [(8, "Terminal")])

    section(ws, 5, "A. PRESSUPOSTOS", span=7)
    label(ws, 6, "ke (desconto)")
    for i in range(5):
        put_formula(ws, 6, 3 + i, f"='{S_PARAMS}'!$C$15", fmt=PCT, ref=True)
    put_formula(ws, 6, 8, f"='{S_PARAMS}'!$C$15", fmt=PCT, ref=True)
    label(ws, 7, "Custo Capital Próprio (ke)")
    for i in range(5):
        put_formula(ws, 7, 3 + i, f"='{S_PARAMS}'!$C$15", fmt=PCT, ref=True)

    section(ws, 8, "B. PROJEÇÃO FCFE (€ mil)", span=7)
    label(ws, 9, "Lucro Líquido (NI)")
    for i in range(5):
        put_input(ws, 9, 3 + i, fmt=EURMIL, source=("← M6: DR" if i == 0 else None))
    label(ws, 10, "(−) CapEx Líquido (CapEx−D&A)")
    for i in range(5):
        put_input(ws, 10, 3 + i, fmt=EURMIL, source=("← M6" if i == 0 else None))
    label(ws, 11, "(−) Δ Fundo de Maneio")
    for i, col in enumerate(YEAR_COLS):
        put_formula(ws, 11, 3 + i, f"='{S_DCF}'!{col}22", fmt=EURMIL, ref=True)
    label(ws, 12, "(+) Nova Dívida Líquida")
    for i in range(5):
        put_input(ws, 12, 3 + i, fmt=EURMIL, source=("← M6: Δ dívida fin." if i == 0 else None))
    label(ws, 13, "FCFE", bold=True)
    for i, col in enumerate(YEAR_COLS):
        put_formula(ws, 13, 3 + i, f"={col}9-{col}10-{col}11+{col}12", fmt=EURMIL)
    put_formula(ws, 13, 8, f"=G13*(1+'{S_PARAMS}'!C27)", fmt=EURMIL)

    section(ws, 14, "C. DESCONTO E EQUITY VALUE (€ mil)", span=7)
    label(ws, 15, "Fator de Desconto (ke)")
    for i, col in enumerate(YEAR_COLS):
        put_formula(ws, 15, 3 + i, f"=1/(1+{col}7)^{i+1}", fmt=FACT)
    label(ws, 16, "FCFE Descontado")
    for i, col in enumerate(YEAR_COLS):
        put_formula(ws, 16, 3 + i, f"={col}13*{col}15", fmt=EURMIL)
    label(ws, 17, "Valor Terminal FCFE")
    put_formula(ws, 17, 3, f"=IFERROR(H13/('{S_PARAMS}'!C15-'{S_PARAMS}'!C27),\"\")", fmt=EURMIL)
    label(ws, 18, "Valor Atual do Valor Terminal")
    put_formula(ws, 18, 3, "=C17*G15", fmt=EURMIL)
    label(ws, 19, "Soma FCFE Descontados (1-n)")
    put_formula(ws, 19, 3, "=SUM(C16:G16)", fmt=EURMIL)
    label(ws, 20, "Equity Value (FCFE)", bold=True)
    put_formula(ws, 20, 3, "=C18+C19", fmt=EURMIL)
    label(ws, 21, "Valor por Ação (€)")
    put_formula(ws, 21, 3, f"=IFERROR(C20*1000/'{S_PARAMS}'!C37,\"\")", fmt=EUR2)


# ── Stress Tests (grelha dinâmica) ───────────────────────────────────────────────
def _stress(ws):
    ws.column_dimensions["B"].width = 16
    for col in YEAR_COLS:
        ws.column_dimensions[col].width = 14
    title(ws, "📉  ANÁLISE DE SENSIBILIDADE — STRESS TESTS (dinâmico)", span=7)
    label(ws, 3, "TABELA — Equity Value DCF/FCFF em função de WACC e g terminal (Δ p.p.)")
    head(ws, 4, [(2, "Δ WACC \\ Δg →")] + [(3 + i, lbl) for i, lbl in
         enumerate(["-1.0%", "-0.5%", "+0.0%", "+0.5%", "+1.0%"])])
    g_off = [-0.01, -0.005, 0.0, 0.005, 0.01]
    w_off = [-0.02, -0.01, 0.0, 0.01, 0.02]
    # cabeçalhos de offset (numéricos, para a fórmula)
    for j, g in enumerate(g_off):
        c = ws.cell(4, 3 + j); c.value = ["-1.0%", "-0.5%", "+0.0%", "+0.5%", "+1.0%"][j]
    for i, w in enumerate(w_off):
        c = ws.cell(5 + i, 2, ["-2.0%", "-1.0%", "+0.0%", "+1.0%", "+2.0%"][i])
        c.font = F_HEAD; c.fill = FILL_HEAD; c.alignment = ALIGN_C; c.border = BORDER

    base_w = f"'{S_PARAMS}'!$C$23"
    base_g = f"'{S_PARAMS}'!$C$27"
    nd = f"'{S_PARAMS}'!$C$36"
    fcff = [f"'{S_DCF}'!${c}$23" for c in YEAR_COLS]  # FCFF anos 1-5

    for i, w in enumerate(w_off):
        for j, g in enumerate(g_off):
            r = 5 + i
            cc = 3 + j
            W = f"({base_w}+({w}))"
            G = f"({base_g}+({g}))"
            pv = "+".join(f"{fcff[t]}/(1+{W})^{t+1}" for t in range(5))
            tv = f"({fcff[4]}*(1+{G})/({W}-{G}))/(1+{W})^5"
            formula = f'=IFERROR(({pv}+{tv})-{nd},"")'
            cell = ws.cell(r, cc, formula)
            cell.font = F_CALC
            cell.number_format = EURMIL
            cell.alignment = ALIGN_R
            cell.border = BORDER
    label(ws, 11, "Cada célula recalcula o Equity Value (DCF) ao WACC e g indicados — "
                  "totalmente dinâmico, sem valores fixos.", note=True)


# ── Síntese ─────────────────────────────────────────────────────────────────────
def _syn(ws):
    ws.column_dimensions["B"].width = 42
    for col in ["C", "D", "E"]:
        ws.column_dimensions[col].width = 18
    title(ws, "🏆  SÍNTESE — COMPARAÇÃO E PREÇO DE NEGOCIAÇÃO", span=4)
    head(ws, 3, [(2, "Metodologia"), (3, "Equity Value (€ mil)"), (4, "Peso (%)"),
                 (5, "Contribuição (€ mil)")])
    label(ws, 4, "1. DCF — FCFF (Obrigatório)")
    put_formula(ws, 4, 3, f"='{S_DCF}'!C35", fmt=EURMIL, ref=True)
    put_input(ws, 4, 4, fmt=PCT, source="peso (input)")
    put_formula(ws, 4, 5, "=C4*D4", fmt=EURMIL)
    label(ws, 5, "2. Múltiplos (mediana EV/EBITDA, EV/EBIT, EV/Sales)")
    put_formula(ws, 5, 3, f"='{S_MULT}'!D17", fmt=EURMIL, ref=True)
    put_input(ws, 5, 4, fmt=PCT, source="peso (input)")
    put_formula(ws, 5, 5, "=C5*D5", fmt=EURMIL)
    label(ws, 6, "3. FCFE (ótica do acionista)")
    put_formula(ws, 6, 3, f"='{S_FCFE}'!C20", fmt=EURMIL, ref=True)
    put_input(ws, 6, 4, fmt=PCT, source="peso (input)")
    put_formula(ws, 6, 5, "=C6*D6", fmt=EURMIL)
    label(ws, 7, "VALOR PONDERADO", bold=True)
    put_formula(ws, 7, 4, "=SUM(D4:D6)", fmt=PCT)
    put_formula(ws, 7, 5, '=IFERROR(SUMPRODUCT(C4:C6,D4:D6)/SUM(D4:D6),"")', fmt=EURMIL)

    section(ws, 9, "DECISÃO FINAL — PREÇO MÍNIMO DE NEGOCIAÇÃO", span=4)
    label(ws, 10, "Valor Ponderado de Referência (€ mil)")
    put_formula(ws, 10, 3, "=E7", fmt=EURMIL)
    label(ws, 11, "Desconto de Negociação (%)")
    put_input(ws, 11, 3, fmt=PCT, source="ex.: −10% (margem negocial)")
    label(ws, 12, "Preço Mínimo (€ mil)", bold=True)
    put_formula(ws, 12, 3, "=C10*(1+C11)", fmt=EURMIL)
    label(ws, 13, "Preço Mínimo por Ação (€)", bold=True)
    put_formula(ws, 13, 3, f"=IFERROR(C12*1000/'{S_PARAMS}'!C37,\"\")", fmt=EUR2)


if __name__ == "__main__":
    p = build()
    print("Gerado:", p)
